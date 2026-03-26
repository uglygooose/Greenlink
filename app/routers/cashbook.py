# app/routers/cashbook.py
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
from typing import List, Optional
from pydantic import BaseModel
from io import BytesIO
import csv
from io import StringIO
import json
import re
from decimal import Decimal, ROUND_HALF_UP

from app.auth import get_db, get_current_user
from app.audit import record_audit_event
from app.models import Booking, TeeTime, BookingStatus, LedgerEntry, LedgerEntryMeta, DayClose, AccountingSetting, User, UserRole, ClubSetting, ProShopSale
from app.fee_models import FeeCategory
from app.club_config import invalidate_club_config_cache
from app.observability import log_event
from app.services.cashbook_service import build_daily_journal_lines
from app.services.finance_semantics_service import (
    build_finance_semantics_metadata,
    build_ledger_entry_finance_state,
    get_export_mapping_status,
    summarize_ledger_finance_states,
)
from app.tenancy import get_active_club_id
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/cashbook", tags=["cashbook"], dependencies=[Depends(get_active_club_id)])

_PRO_SHOP_EXPORTS_SETTING_KEY = "cashbook_pro_shop_exports"


def verify_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


def _request_id(request: Request | None) -> str | None:
    if request is None:
        return None
    return str(getattr(getattr(request, "state", None), "request_id", "") or "").strip() or None


def _audit_cashbook_event(
    db: Session,
    request: Request | None,
    actor: User | None,
    action: str,
    entity_type: str,
    *,
    entity_id: str | int | None = None,
    payload: dict | None = None,
) -> None:
    record_audit_event(
        db,
        action=action,
        entity_type=entity_type,
        actor_user_id=int(getattr(actor, "id", 0) or 0) or None,
        entity_id=entity_id,
        payload=payload,
        request_id=_request_id(request),
        club_id=int(getattr(db, "info", {}).get("club_id") or 0) or None,
    )

class PaymentRecord(BaseModel):
    """Individual payment record"""
    period: str  # DD/MM/YYYY
    date: str  # DD/MM/YYYY
    gdc: str  # Golfer identifier
    account_number: str  # Account number from booking
    reference: str  # Reference code
    description: str  # Description of payment
    amount: float  # Payment amount
    tax_type: int  # 0=No tax, 1=Tax
    tax_amount: float  # Tax amount
    open_item: str  # Open item code
    projects_code: str  # Projects code (blank for now)
    contra_account: str  # Contra account (GL Account)
    exchange_rate: float  # Exchange rate (default 1)
    bank_exchange_rate: float  # Bank exchange rate (default 1)
    batch_id: int  # Batch identifier
    discount_tax_type: int  # 0=No discount tax, 1=Discount tax
    discount_amount: float  # Discount amount
    home_amount: float  # Home amount (same as amount in single currency)


class DailyPaymentsSummary(BaseModel):
    """Summary of payments collected"""
    date: str
    total_payments: float
    total_tax: float
    transaction_count: int
    records: List[PaymentRecord]


class AccountingSettingsPayload(BaseModel):
    green_fees_gl: Optional[str] = None
    cashbook_contra_gl: Optional[str] = None
    vat_rate: Optional[float] = None
    tax_type: Optional[int] = None
    cashbook_name: Optional[str] = None


def get_accounting_settings(db: Session) -> AccountingSetting:
    club_id = getattr(db, "info", {}).get("club_id")
    if not club_id:
        raise HTTPException(status_code=400, detail="club_id is required")

    settings = db.query(AccountingSetting).filter(AccountingSetting.club_id == int(club_id)).first()
    if settings:
        return settings
    settings = AccountingSetting(club_id=int(club_id))
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


def _upsert_club_setting(db: Session, key: str, value: str) -> None:
    club_id = getattr(db, "info", {}).get("club_id")
    if not club_id:
        raise HTTPException(status_code=400, detail="club_id is required")

    row = db.query(ClubSetting).filter(ClubSetting.club_id == int(club_id), ClubSetting.key == key).first()
    if row:
        row.value = value
        row.updated_at = datetime.utcnow()
    else:
        db.add(ClubSetting(club_id=int(club_id), key=key, value=value))
    invalidate_club_config_cache(int(club_id))


def _get_club_setting(db: Session, key: str) -> Optional[str]:
    club_id = getattr(db, "info", {}).get("club_id")
    if not club_id:
        return None
    row = db.query(ClubSetting).filter(ClubSetting.club_id == int(club_id), ClubSetting.key == key).first()
    if not row or row.value is None:
        return None
    return str(row.value)


def _get_gl_account_reference(db: Session) -> dict:
    raw = _get_club_setting(db, "gl_account_reference")
    if not raw:
        return {"configured": False, "count": 0, "source_file": None, "accounts": []}
    try:
        parsed = json.loads(raw) or {}
    except Exception:
        return {"configured": True, "count": 0, "source_file": None, "accounts": []}

    accounts_raw = parsed.get("accounts") if isinstance(parsed, dict) else []
    accounts: list[dict[str, str]] = []
    seen: set[str] = set()
    if isinstance(accounts_raw, list):
        for row in accounts_raw:
            if not isinstance(row, dict):
                continue
            account = str(row.get("account") or "").strip()
            description = str(row.get("description") or "").strip()
            if not account or not description:
                continue
            key = account.lower()
            if key in seen:
                continue
            seen.add(key)
            accounts.append({"account": account, "description": description})

    return {
        "configured": bool(accounts),
        "count": len(accounts),
        "source_file": str(parsed.get("source_file") or "").strip() or None,
        "accounts": accounts,
    }


def _infer_date_format(sample: str) -> Optional[str]:
    raw = (sample or "").strip()
    if not raw:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        return "YYYY-MM-DD"
    if re.match(r"^\d{2}/\d{2}/\d{4}$", raw):
        return "DD/MM/YYYY"
    if re.match(r"^\d{2}/\d{2}/\d{2}$", raw):
        return "DD/MM/YY"
    if re.match(r"^\d{4}/\d{2}/\d{2}$", raw):
        return "YYYY/MM/DD"
    return None


def _best_header_match(headers: List[str], *needles: str) -> Optional[str]:
    """
    Return the first header that contains all needle fragments (case-insensitive),
    preferring exact-ish matches.
    """
    normalized = [(h, re.sub(r"[^a-z0-9]+", "", (h or "").strip().lower())) for h in headers]
    want = [re.sub(r"[^a-z0-9]+", "", (n or "").strip().lower()) for n in needles if n]
    if not want:
        return None

    # Exact normalized match
    for h, n in normalized:
        if n in want:
            return h

    # Contains all fragments
    for h, n in normalized:
        if all(w in n for w in want):
            return h
    return None


def _build_layout_column_map(headers: List[str]) -> dict:
    return {
        "date": _best_header_match(headers, "date"),
        "reference": _best_header_match(headers, "ref") or _best_header_match(headers, "reference"),
        "description": _best_header_match(headers, "desc") or _best_header_match(headers, "description"),
        "account": _best_header_match(headers, "account") or _best_header_match(headers, "gl"),
        "debit": _best_header_match(headers, "debit"),
        "credit": _best_header_match(headers, "credit"),
        "amount": _best_header_match(headers, "amount") or _best_header_match(headers, "value"),
        "tax_flag": _best_header_match(headers, "tax", "flag") or _best_header_match(headers, "vat", "flag"),
        "tax_type": _best_header_match(headers, "taxtype") or _best_header_match(headers, "tax", "type"),
        "tax_amount": _best_header_match(headers, "taxamount") or _best_header_match(headers, "tax", "amount"),
    }


_DATE_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
_ACCOUNT_RE = re.compile(r"^\d{5,10}$")


def _is_date_cell(value: str) -> bool:
    return bool(_DATE_RE.match(str(value or "").strip()))


def _is_account_cell(value: str) -> bool:
    return bool(_ACCOUNT_RE.match(str(value or "").strip()))


def _is_amount_cell(value: str) -> bool:
    s = str(value or "").strip()
    if not s:
        return False
    # Pastel exports often use dot decimals; require a dot or minus sign to avoid
    # accidentally treating batch numbers (e.g. "13") as an amount column.
    if "." not in s and "-" not in s:
        return False
    try:
        Decimal(s)
        return True
    except Exception:
        return False


def _guess_headerless_layout_map(headers: List[str], sample_row: List[str]) -> dict:
    """
    Best-effort mapping for Pastel Batch->Export files that don't include headers.
    Uses the first row as a template and maps common fields by position/pattern.
    """
    n = len(sample_row)
    if n != len(headers):
        headers = [f"COL_{i+1}" for i in range(n)]

    idx_date = next((i for i, v in enumerate(sample_row) if _is_date_cell(v)), None)
    idx_account = next((i for i, v in enumerate(sample_row) if _is_account_cell(v)), None)
    idx_amount = next((i for i, v in enumerate(sample_row) if _is_amount_cell(v)), None)
    idx_tax_flag = None
    idx_tax_amount = None
    if idx_amount is not None and (idx_amount + 1) < n:
        nxt = str(sample_row[idx_amount + 1] or "").strip()
        if nxt in {"0", "1"}:
            idx_tax_flag = idx_amount + 1
            if (idx_amount + 2) < n and _is_amount_cell(sample_row[idx_amount + 2]):
                idx_tax_amount = idx_amount + 2

    # Description/reference are typically free-text; pick the first meaningful text after account.
    idx_desc = None
    idx_ref = None
    if idx_account is not None:
        for i in range(idx_account + 1, n):
            v = str(sample_row[i] or "").strip()
            if not v:
                continue
            if idx_desc is None and any(c.isalpha() for c in v):
                idx_desc = i
                continue
            if idx_desc is not None and idx_ref is None and any(c.isalpha() for c in v):
                idx_ref = i
                break

    # Fallbacks for the common Pastel export shape shown by clubs (18 columns):
    # 1=batch, 2=date, 3=journal, 4=account, 5=desc, 6=ref/type, 7=amount
    if n >= 7:
        idx_date = idx_date if idx_date is not None else 1
        idx_account = idx_account if idx_account is not None else 3
        idx_desc = idx_desc if idx_desc is not None else 4
        idx_ref = idx_ref if idx_ref is not None else 5
        idx_amount = idx_amount if idx_amount is not None else 6
        if idx_tax_flag is None and n >= 9:
            idx_tax_flag = 7
        if idx_tax_amount is None and n >= 9:
            idx_tax_amount = 8

    # Tax type is often a short code like GOV01; it tends to appear later.
    idx_tax_type = None
    for i in range(n - 1, -1, -1):
        v = str(sample_row[i] or "").strip()
        if len(v) in {4, 5, 6, 7} and any(c.isalpha() for c in v) and any(c.isdigit() for c in v):
            idx_tax_type = i
            break

    def _h(i: int | None) -> str | None:
        if i is None:
            return None
        if 0 <= i < len(headers):
            return headers[i]
        return None

    # Pastel commonly displays "Reference" before "Description". In many exports, the first
    # text column after account is the on-screen Reference, and the next is Description.
    return {
        "date": _h(idx_date),
        "account": _h(idx_account),
        "reference": _h(idx_desc),
        "description": _h(idx_ref),
        "amount": _h(idx_amount),
        "tax_flag": _h(idx_tax_flag),
        "tax_type": _h(idx_tax_type),
        "debit": None,
        "credit": None,
        "tax_amount": _h(idx_tax_amount),
    }


def _looks_like_header_row(row: List[str]) -> bool:
    tokens = ("date", "ref", "reference", "desc", "description", "account", "debit", "credit", "amount", "tax")
    normalized = ["".join(ch.lower() for ch in str(c or "") if ch.isalnum()) for c in row]
    return any(any(t in cell for t in tokens) for cell in normalized)


@router.get("/pastel-layout")
def get_pastel_layout(
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    raw = _get_club_setting(db, "pastel_journal_layout")
    if not raw:
        return {"configured": False}
    try:
        return {"configured": True, "layout": json.loads(raw)}
    except Exception:
        return {"configured": True, "layout_raw": raw}


@router.post("/pastel-layout")
async def upload_pastel_layout(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    """
    Upload a Pastel-exported journal CSV (Batch -> Export) so we can match the import layout exactly.
    Stores the parsed layout in ClubSetting key: pastel_journal_layout
    """
    try:
        data = await file.read()
        text = data.decode("utf-8-sig", errors="replace")
        sample = "\n".join(text.splitlines()[:10])

        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(sample, delimiters=[",", ";", "|", "\t"])
        except Exception:
            dialect = csv.excel
        try:
            has_header = bool(sniffer.has_header(sample))
        except Exception:
            has_header = True

        reader = csv.reader(StringIO(text), dialect)
        rows: list[list[str]] = []
        for row in reader:
            if row and any(str(cell or "").strip() for cell in row):
                rows.append([str(cell or "").strip() for cell in row])
            if len(rows) >= 25:
                break

        if not rows:
            raise HTTPException(status_code=400, detail="Empty CSV file")

        # csv.Sniffer.has_header is often wrong for Pastel exports. Override if the
        # "header" row contains obvious data patterns (date/account/amount) and no
        # header-like tokens.
        if has_header:
            first_row = rows[0]
            if not _looks_like_header_row(first_row) and any(_is_date_cell(v) for v in first_row):
                has_header = False

        sample_rows: list[list[str]] = []
        if has_header:
            headers = rows[0]
            sample_rows = rows[1:11] if len(rows) > 1 else []
            first_data = sample_rows[0] if sample_rows else None
            column_map = _build_layout_column_map(headers)
        else:
            # Headerless Pastel exports: treat the first row as a template and build a positional map.
            first_data = rows[0]
            headers = [f"COL_{i+1}" for i in range(len(first_data))]
            column_map = _guess_headerless_layout_map(headers, first_data)
            sample_rows = rows[:10]

        date_col = _best_header_match(headers, "date")
        date_format = None
        if first_data and column_map.get("date") and column_map.get("date") in headers:
            idx = headers.index(column_map.get("date"))
            if idx < len(first_data):
                date_format = _infer_date_format(first_data[idx])

        # Extra inference to help per-client setups.
        inferred = {}
        try:
            amount_hdr = column_map.get("amount")
            account_hdr = column_map.get("account")
            tax_type_hdr = column_map.get("tax_type")
            tax_flag_hdr = column_map.get("tax_flag")
            tax_amount_hdr = column_map.get("tax_amount")

            amount_mirrors: list[str] = []
            account_digits_only = False
            observed_tax_types: list[str] = []
            inferred_amount_sign: str | None = None

            if first_data and account_hdr and account_hdr in headers:
                aidx = headers.index(account_hdr)
                if aidx < len(first_data):
                    account_digits_only = _is_account_cell(first_data[aidx])

            if first_data and amount_hdr and amount_hdr in headers:
                midx = headers.index(amount_hdr)
                if midx < len(first_data):
                    amount_val = str(first_data[midx] or "").strip()
                    for i, v in enumerate(first_data):
                        if i == midx:
                            continue
                        if str(v or "").strip() == amount_val and _is_amount_cell(v) and _is_amount_cell(amount_val):
                            amount_mirrors.append(headers[i])

            if tax_type_hdr and tax_type_hdr in headers and sample_rows:
                tidx = headers.index(tax_type_hdr)
                seen = set()
                for r in sample_rows:
                    if tidx >= len(r):
                        continue
                    v = str(r[tidx] or "").strip()
                    if not v:
                        continue
                    if v not in seen:
                        seen.add(v)
                        observed_tax_types.append(v)

            if amount_hdr and tax_flag_hdr and amount_hdr in headers and tax_flag_hdr in headers and sample_rows:
                midx = headers.index(amount_hdr)
                fidx = headers.index(tax_flag_hdr)
                for r in sample_rows:
                    if midx >= len(r) or fidx >= len(r):
                        continue
                    if str(r[fidx] or "").strip() != "1":
                        continue
                    try:
                        amt = Decimal(str(r[midx]).strip())
                    except Exception:
                        continue
                    inferred_amount_sign = "debit_positive" if amt < 0 else "debit_negative"
                    break

            inferred = {
                "amount_mirrors": amount_mirrors,
                "account_digits_only": account_digits_only,
                "observed_tax_types": observed_tax_types,
                "inferred_amount_sign": inferred_amount_sign,
                "has_tax_flag": bool(tax_flag_hdr and tax_flag_hdr in headers),
                "has_tax_amount": bool(tax_amount_hdr and tax_amount_hdr in headers),
            }
        except Exception:
            inferred = {}

        layout = {
            "delimiter": getattr(dialect, "delimiter", ","),
            "has_header": has_header,
            "columns": headers,
            "date_format": date_format,
            "column_map": column_map,
            "template_row": first_data,
            "sample_rows": sample_rows,
            "inferred": inferred,
            "uploaded_at": datetime.utcnow().isoformat(),
            "filename": file.filename or None,
        }

        _upsert_club_setting(db, "pastel_journal_layout", json.dumps(layout))
        db.commit()

        return {"status": "success", "layout": layout}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[PASTEL] Layout upload failed: {str(e)[:240]}")
        raise HTTPException(status_code=500, detail="Failed to process Pastel layout CSV")


class PastelJournalMappingsPayload(BaseModel):
    vat_output_gl: Optional[str] = None
    debit_gl: Optional[dict] = None  # e.g. {"CARD":"8400/000","CASH":"8100/000","EFT":"8410/000","ONLINE":"9450/000","ACCOUNT":"1100/000"}
    tax_type: Optional[str] = None   # Optional (depends on Pastel import layout)
    amount_sign: Optional[str] = None  # "debit_positive" (default) | "debit_negative"


@router.get("/pastel-mappings")
def get_pastel_mappings(
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    raw = _get_club_setting(db, "pastel_journal_mappings")
    if not raw:
        return {"configured": False}
    try:
        return {"configured": True, "mappings": json.loads(raw)}
    except Exception:
        return {"configured": True, "mappings_raw": raw}


@router.put("/pastel-mappings")
def update_pastel_mappings(
    payload: PastelJournalMappingsPayload,
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    raw = _get_club_setting(db, "pastel_journal_mappings")
    current = {}
    if raw:
        try:
            current = json.loads(raw) or {}
        except Exception:
            current = {}

    if payload.vat_output_gl is not None:
        current["vat_output_gl"] = (payload.vat_output_gl or "").strip() or None

    if payload.tax_type is not None:
        current["tax_type"] = (payload.tax_type or "").strip() or None

    if payload.debit_gl is not None:
        debit_gl = {}
        for k, v in (payload.debit_gl or {}).items():
            key = str(k or "").strip().upper()
            val = str(v or "").strip()
            if not key:
                continue
            debit_gl[key] = val or None
        current["debit_gl"] = debit_gl

    if payload.amount_sign is not None:
        val = str(payload.amount_sign or "").strip().lower()
        if val and val not in {"debit_positive", "debit_negative"}:
            raise HTTPException(status_code=400, detail="amount_sign must be 'debit_positive' or 'debit_negative'")
        current["amount_sign"] = val or None

    current["updated_at"] = datetime.utcnow().isoformat()

    _upsert_club_setting(db, "pastel_journal_mappings", json.dumps(current))
    db.commit()

    return {"status": "success", "mappings": current}


def get_active_completed_bookings(db: Session, target_date: Optional[date] = None) -> List:
    """
    Get all bookings that were paid on a specific date.

    Accounting rule: the transaction date is the payment date (when the ledger entry was created),
    not the tee time date.
    """
    if target_date is None:
        target_date = date.today()
    
    bookings = (
        db.query(Booking)
        .join(LedgerEntry, LedgerEntry.booking_id == Booking.id)
        .filter(
            func.date(LedgerEntry.created_at) == target_date,
            Booking.status.in_([BookingStatus.checked_in, BookingStatus.completed]),
        )
        .distinct()
        .all()
    )
    
    return bookings


def sanitize_gl(value: Optional[str]) -> str:
    if not value:
        return ""
    return value.replace("/", "").replace(" ", "").replace("-", "")


def format_amount(value: float) -> str:
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return text if text else "0"


def _q2(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _money_str(value: Decimal) -> str:
    v = _q2(value)
    if v == 0:
        return ""
    return f"{v:.2f}"


def _money_str_zero(value: Decimal) -> str:
    v = _q2(value)
    if v == 0:
        return "0"
    return f"{v:.2f}"


def _format_gl_for_layout(gl: str, layout: dict) -> str:
    """
    Pastel exports sometimes use digits-only account codes (e.g. 9500000) even
    if the UI shows 9500/000. When the uploaded template indicates digits-only,
    we sanitize GLs accordingly.
    """
    raw = str(gl or "").strip()
    inferred = (layout or {}).get("inferred") or {}
    if inferred.get("account_digits_only"):
        return re.sub(r"[^0-9]", "", raw)
    return raw


def _clean_text(value: str, max_len: int = 60) -> str:
    raw = (value or "").strip()
    raw = re.sub(r"[^A-Za-z0-9 _\\-]", " ", raw)
    raw = re.sub(r"\\s+", " ", raw).strip()
    if max_len and len(raw) > max_len:
        raw = raw[:max_len].strip()
    return raw


def _format_date_for_layout(d: date, date_format: Optional[str]) -> str:
    fmt = (date_format or "").strip().upper()
    if fmt == "DD/MM/YYYY":
        return d.strftime("%d/%m/%Y")
    if fmt == "DD/MM/YY":
        return d.strftime("%d/%m/%y")
    if fmt == "YYYY/MM/DD":
        return d.strftime("%Y/%m/%d")
    if fmt == "YYYY-MM-DD":
        return d.strftime("%Y-%m-%d")
    # Default
    return d.strftime("%Y-%m-%d")


def _require_pastel_layout(db: Session) -> dict:
    raw = _get_club_setting(db, "pastel_journal_layout")
    if not raw:
        raise HTTPException(status_code=400, detail="Pastel journal layout is not configured. Upload the Pastel-exported CSV first.")
    try:
        layout = json.loads(raw) or {}
    except Exception:
        raise HTTPException(status_code=500, detail="Pastel journal layout is invalid. Re-upload the layout CSV.")

    cols = layout.get("columns") or []
    if not isinstance(cols, list) or not cols:
        raise HTTPException(status_code=500, detail="Pastel journal layout is missing columns. Re-upload the layout CSV.")
    return layout


def _require_pastel_mappings(db: Session) -> dict:
    raw = _get_club_setting(db, "pastel_journal_mappings")
    if not raw:
        raise HTTPException(status_code=400, detail="Pastel mappings are not configured. Save VAT + debit GL mappings first.")
    try:
        mappings = json.loads(raw) or {}
    except Exception:
        raise HTTPException(status_code=500, detail="Pastel mappings are invalid. Re-save mappings.")

    vat_gl = (mappings.get("vat_output_gl") or "").strip()
    if not vat_gl:
        raise HTTPException(
            status_code=400,
            detail="Missing Output VAT GL account in Pastel mappings. Set it in Admin → Cashbook → Output VAT GL Account, then click 'Save Pastel Mappings'.",
        )

    debit_gl = mappings.get("debit_gl") or {}
    if not isinstance(debit_gl, dict):
        debit_gl = {}
    mappings["vat_output_gl"] = vat_gl
    mappings["debit_gl"] = {str(k or "").strip().upper(): str(v or "").strip() for k, v in debit_gl.items()}
    return mappings


def _parse_ymd_date(raw_value: Optional[str], *, field_name: str) -> date:
    if raw_value:
        try:
            return datetime.strptime(raw_value, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid {field_name} format. Use YYYY-MM-DD")
    return date.today()


def _normalize_payment_method_for_journal(raw_value: Optional[str]) -> str:
    value = str(raw_value or "").strip().lower()
    if value in {"cash", "card", "eft", "online", "account"}:
        return value.upper()
    if value in {"credit_card", "debit_card", "swipe", "tap", "other", "unknown"}:
        return "CARD"
    return "CARD"


def _pro_shop_tax_split(gross_amount: Decimal, explicit_tax: Decimal, vat_rate: Decimal) -> tuple[Decimal, Decimal]:
    gross = _q2(gross_amount)
    tax = _q2(explicit_tax)
    if tax > 0 and tax <= gross:
        vat_amount = tax
    elif vat_rate > 0:
        vat_amount = _q2(gross * (vat_rate / (Decimal("1") + vat_rate)))
    else:
        vat_amount = Decimal("0.00")
    net_amount = _q2(gross - vat_amount)
    return net_amount, vat_amount


def _get_pro_shop_exports_registry(db: Session) -> dict:
    raw = _get_club_setting(db, _PRO_SHOP_EXPORTS_SETTING_KEY)
    if not raw:
        return {}
    try:
        payload = json.loads(raw)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _set_pro_shop_exports_registry(db: Session, registry: dict) -> None:
    normalized = registry if isinstance(registry, dict) else {}
    _upsert_club_setting(db, _PRO_SHOP_EXPORTS_SETTING_KEY, json.dumps(normalized))


def _render_pastel_journal_csv(
    *,
    layout: dict,
    mappings: dict,
    lines: List[dict],
    target_date: date,
    batch_ref: str,
    vat_rate: Decimal,
) -> str:
    columns = layout.get("columns") or []
    column_map = layout.get("column_map") or {}
    header_to_idx = {str(h): i for i, h in enumerate(columns)}

    has_debit = bool(column_map.get("debit")) and column_map.get("debit") in header_to_idx
    has_credit = bool(column_map.get("credit")) and column_map.get("credit") in header_to_idx
    has_amount = bool(column_map.get("amount")) and column_map.get("amount") in header_to_idx

    for key in ["date", "account"]:
        header = column_map.get(key)
        if not header or header not in header_to_idx:
            raise HTTPException(status_code=500, detail=f"Pastel layout missing required column for '{key}'. Re-upload layout CSV.")

    if not ((has_debit and has_credit) or has_amount):
        raise HTTPException(
            status_code=500,
            detail="Pastel layout must include Debit/Credit columns or an Amount column. Re-upload a Batch->Export sample that matches your import layout.",
        )

    date_format = layout.get("date_format")
    date_value = _format_date_for_layout(target_date, date_format)
    delimiter = layout.get("delimiter") or ","
    has_header = bool(layout.get("has_header", True))
    template_row = layout.get("template_row")
    has_template = isinstance(template_row, list) and len(template_row) == len(columns)
    sample_rows = layout.get("sample_rows") or []
    has_samples = isinstance(sample_rows, list) and bool(sample_rows)

    constant_template: List[Optional[str]] = [None for _ in columns]
    if has_template and has_samples:
        try:
            rows_for_constants = [r for r in sample_rows if isinstance(r, list) and len(r) == len(columns)]
            if rows_for_constants:
                for i in range(len(columns)):
                    vals = [str(r[i] or "").strip() for r in rows_for_constants]
                    if not vals:
                        continue
                    first = vals[0]
                    if first and all(v == first for v in vals[1:]):
                        constant_template[i] = str(template_row[i] or "")
        except Exception:
            constant_template = [None for _ in columns]

    def _base_row() -> list:
        if has_template and any(v is not None for v in constant_template):
            row = ["" for _ in columns]
            for i, v in enumerate(constant_template):
                if v is not None and i < len(row):
                    row[i] = v
            return row
        if has_template:
            return [str(x or "") for x in list(template_row)]
        return ["" for _ in columns]

    def _set(row: list, key: str, value: str) -> None:
        header = column_map.get(key)
        if not header:
            return
        idx = header_to_idx.get(header)
        if idx is None:
            return
        row[idx] = value

    amount_mirror_headers: List[str] = []
    try:
        amount_mirror_headers = list(((layout.get("inferred") or {}).get("amount_mirrors") or []))
    except Exception:
        amount_mirror_headers = []

    def _set_amount_mirrors(row: list, value: str) -> None:
        for h in amount_mirror_headers:
            idx = header_to_idx.get(str(h))
            if idx is None:
                continue
            if idx < len(row):
                row[idx] = value

    out = StringIO(newline="")
    writer = csv.writer(out, delimiter=delimiter, lineterminator="\r\n")
    if has_header:
        writer.writerow(columns)

    tax_type_sales = str((mappings.get("tax_type") or "")).strip()
    tax_type_sales_is_numeric = bool(tax_type_sales) and tax_type_sales.isdigit()
    amount_sign = str((mappings.get("amount_sign") or "")).strip().lower() or "debit_positive"
    vat_output_gl = str((mappings.get("vat_output_gl") or "")).strip()
    vat_output_gl_fmt = _format_gl_for_layout(vat_output_gl, layout) if vat_output_gl else ""

    for line in lines:
        row = _base_row()
        _set(row, "date", date_value)

        if column_map.get("reference") and column_map.get("reference") in header_to_idx:
            _set(row, "reference", batch_ref if has_header else str(line.get("ref") or batch_ref))
        if column_map.get("description") and column_map.get("description") in header_to_idx:
            if has_template:
                d_idx = header_to_idx[column_map.get("description")]
                existing = str(row[d_idx] or "").strip() if d_idx < len(row) else ""
                if not (existing and len(existing) <= 20):
                    _set(row, "description", str(line["desc"]))
            else:
                _set(row, "description", str(line["desc"]))

        account_value = _format_gl_for_layout(str(line["account"]), layout)
        _set(row, "account", account_value)

        is_debit_line = bool(line["debit"] and _q2(line["debit"]) != 0)
        is_credit_line = bool((not is_debit_line) and line["credit"] and _q2(line["credit"]) != 0)
        is_vat_control_line = bool(vat_output_gl_fmt) and account_value.strip() == vat_output_gl_fmt
        is_revenue_line = bool(is_credit_line and not is_vat_control_line)
        apply_sales_tax = bool(is_revenue_line and tax_type_sales)

        amount = None
        if has_amount:
            amount = Decimal("0.00")
            if is_debit_line:
                amount = _q2(line["debit"])
                if amount_sign == "debit_negative":
                    amount = -amount
            else:
                amount = _q2(line["credit"])
                if amount_sign == "debit_positive":
                    amount = -amount
                else:
                    amount = +amount
            amount_str = _money_str(amount)
            _set(row, "amount", amount_str)
            _set_amount_mirrors(row, amount_str)
            if has_debit:
                _set(row, "debit", "")
            if has_credit:
                _set(row, "credit", "")
        else:
            if is_debit_line:
                _set(row, "debit", _money_str(line["debit"]))
                _set(row, "credit", "")
            else:
                _set(row, "debit", "")
                _set(row, "credit", _money_str(line["credit"]))

        if column_map.get("tax_type") and column_map.get("tax_type") in header_to_idx:
            if apply_sales_tax and tax_type_sales:
                _set(row, "tax_type", tax_type_sales)
            else:
                _set(row, "tax_type", "00" if tax_type_sales_is_numeric else "")

        if column_map.get("tax_flag") and column_map.get("tax_flag") in header_to_idx:
            _set(row, "tax_flag", "1" if apply_sales_tax else "0")

        if column_map.get("tax_amount") and column_map.get("tax_amount") in header_to_idx:
            if apply_sales_tax:
                vat_amt = _q2(_to_decimal(line.get("vat", 0) or 0))
                if vat_amt == 0:
                    base_net = _q2(line["credit"]) if is_credit_line else _q2(line["debit"])
                    vat_amt = _q2(base_net * vat_rate) if vat_rate > 0 else Decimal("0.00")
                if amount is not None and amount < 0:
                    vat_amt = -vat_amt
                _set(row, "tax_amount", _money_str_zero(vat_amt))
            else:
                _set(row, "tax_amount", "0")

        writer.writerow(row)

    return out.getvalue()


def create_payment_record(
    booking: Booking,
    settings: AccountingSetting,
    payment_date: Optional[date] = None,
    batch_id: int = 1,
) -> PaymentRecord:
    """Convert a booking to a payment record"""
    tee_time = booking.tee_time
    fee_category = booking.fee_category_id
    
    # Get fee details if available
    amount = float(booking.price or 0.0)
    
    # Create reference from booking ID and player
    reference = f"BK{booking.id:05d}"  # 5 chars max for reference
    description = f"Golf Fee - {booking.player_name}"
    
    # Format dates
    if payment_date:
        date_obj = datetime.combine(payment_date, datetime.min.time())
    else:
        date_obj = tee_time.tee_time if tee_time else datetime.now()
    date_str = date_obj.strftime("%d/%m/%Y")
    
    # Period: month number (1-12)
    period_num = date_obj.month
    
    # Calculate tax (defaults to 15% VAT for SA)
    tax_type = settings.tax_type if settings else 1
    tax_rate = settings.vat_rate if settings else 0.15
    if tax_type and tax_rate:
        tax_amount = round(amount * tax_rate / (1 + tax_rate), 2)  # Extract tax from inclusive price
    else:
        tax_amount = 0.0
    
    # Account number: GL account from settings
    account_number = sanitize_gl(settings.green_fees_gl if settings else "1000-000")
    
    # GDC: Use "G" for General Ledger
    gdc = "G"
    
    # Contra account: cashbook bank account (no slashes/spaces)
    contra_account = sanitize_gl(settings.cashbook_contra_gl if settings else "8400/000")
    
    return PaymentRecord(
        period=str(period_num),  # Month number as string
        date=date_str,
        gdc=gdc,  # General Ledger
        account_number=account_number,  # 7 chars max
        reference=reference,
        description=description,
        amount=amount,
        tax_type=tax_type,
        tax_amount=tax_amount,
        open_item=" ",
        projects_code="     ",
        contra_account=contra_account,  # No slashes
        exchange_rate=1,
        bank_exchange_rate=1,
        batch_id=batch_id,
        discount_tax_type=0,
        discount_amount=0,
        home_amount=amount
    )


def create_excel_workbook(payments: List[PaymentRecord], date_str: str) -> BytesIO:
    """Create an Excel workbook with payment records in the specified format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Define headers matching the Excel template
    headers = [
        "Period", "Date", "GDC", "Account Number", "Reference", "Description",
        "Amount", "Tax Type", "Tax Amount", "Open Item", "Projects Code",
        "Contra Account", "Exchange Rate", "Bank Exchange Rate", "Batch ID",
        "Discount Tax Type", "Discount Amount", "Home Amount"
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    data_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    
    for row_idx, payment in enumerate(payments, start=2):
        ws.cell(row=row_idx, column=1).value = payment.period
        ws.cell(row=row_idx, column=2).value = payment.date
        ws.cell(row=row_idx, column=3).value = payment.gdc
        ws.cell(row=row_idx, column=4).value = payment.account_number
        ws.cell(row=row_idx, column=5).value = payment.reference
        ws.cell(row=row_idx, column=6).value = payment.description
        ws.cell(row=row_idx, column=7).value = payment.amount
        ws.cell(row=row_idx, column=8).value = payment.tax_type
        ws.cell(row=row_idx, column=9).value = payment.tax_amount
        ws.cell(row=row_idx, column=10).value = payment.open_item
        ws.cell(row=row_idx, column=11).value = payment.projects_code
        ws.cell(row=row_idx, column=12).value = payment.contra_account
        ws.cell(row=row_idx, column=13).value = payment.exchange_rate
        ws.cell(row=row_idx, column=14).value = payment.bank_exchange_rate
        ws.cell(row=row_idx, column=15).value = payment.batch_id
        ws.cell(row=row_idx, column=16).value = payment.discount_tax_type
        ws.cell(row=row_idx, column=17).value = payment.discount_amount
        ws.cell(row=row_idx, column=18).value = payment.home_amount
        
        # Apply styling to data rows
        for col_idx in range(1, 19):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # Number formatting for amount columns
            if col_idx in [7, 9, 13, 14, 17, 18]:  # Amount columns
                cell.number_format = '#,##0.00'
                cell.alignment = number_alignment
            elif col_idx in [8, 15, 16]:  # Type columns
                cell.number_format = '0'
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment
    
    # Adjust column widths
    column_widths = [12, 12, 15, 16, 12, 20, 12, 10, 12, 12, 14, 16, 12, 16, 10, 16, 16, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def create_csv_content(payments: List[PaymentRecord]) -> StringIO:
    output = StringIO()
    writer = csv.writer(output)
    for p in payments:
        writer.writerow([
            p.period,
            p.date,
            p.gdc,
            p.account_number,
            p.reference,
            p.description,
            format_amount(p.amount),
            p.tax_type,
            format_amount(p.tax_amount),
            p.open_item,
            p.projects_code,
            p.contra_account,
            format_amount(p.exchange_rate),
            format_amount(p.bank_exchange_rate),
            p.batch_id,
            p.discount_tax_type,
            format_amount(p.discount_amount),
            format_amount(p.home_amount)
        ])
    output.seek(0)
    return output


@router.get("/daily-summary")
def get_daily_summary(
    summary_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
) -> DailyPaymentsSummary:
    """Get summary of all payments collected for a specific day"""
    
    # Parse date
    if summary_date:
        try:
            target_date = datetime.strptime(summary_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()
    
    # Get bookings
    try:
        bookings = get_active_completed_bookings(db, target_date)
    except Exception as e:
        # Database not available, return empty summary
        return DailyPaymentsSummary(
            date=target_date.strftime("%Y-%m-%d"),
            total_payments=0.0,
            total_tax=0.0,
            transaction_count=0,
            records=[]
        )
    
    # Convert to payment records
    settings = get_accounting_settings(db)
    records = [create_payment_record(booking, settings, payment_date=target_date) for booking in bookings]
    
    # Calculate totals
    total_payments = sum(r.amount for r in records)
    total_tax = sum(r.tax_amount for r in records)
    
    return DailyPaymentsSummary(
        date=target_date.strftime("%Y-%m-%d"),
        total_payments=total_payments,
        total_tax=total_tax,
        transaction_count=len(records),
        records=records
    )


@router.get("/pro-shop-summary")
def get_pro_shop_summary(
    summary_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    target_date = _parse_ymd_date(summary_date, field_name="summary_date")
    club_id = int(getattr(db, "info", {}).get("club_id") or 0)
    if club_id <= 0:
        raise HTTPException(status_code=400, detail="club_id is required")

    try:
        sales = (
            db.query(ProShopSale)
            .filter(ProShopSale.club_id == club_id, func.date(ProShopSale.sold_at) == target_date)
            .order_by(ProShopSale.sold_at.asc(), ProShopSale.id.asc())
            .all()
        )
    except Exception:
        raise HTTPException(status_code=503, detail="Database connection unavailable")

    settings = get_accounting_settings(db)
    try:
        vat_rate = Decimal(str(settings.vat_rate if getattr(settings, "vat_rate", None) is not None else 0.15))
    except Exception:
        vat_rate = Decimal("0.15")
    if vat_rate < 0:
        vat_rate = Decimal("0")

    total_payments = Decimal("0.00")
    total_tax = Decimal("0.00")
    method_totals: dict[str, Decimal] = {}
    method_counts: dict[str, int] = {}
    records: list[dict] = []

    for sale in sales:
        gross = _q2(_to_decimal(getattr(sale, "total", 0) or 0))
        if gross <= 0:
            continue
        method = _normalize_payment_method_for_journal(getattr(sale, "payment_method", None))
        explicit_tax = _q2(_to_decimal(getattr(sale, "tax", 0) or 0))
        _, vat_amount = _pro_shop_tax_split(gross, explicit_tax, vat_rate)

        method_totals[method] = _q2(method_totals.get(method, Decimal("0.00")) + gross)
        method_counts[method] = int(method_counts.get(method, 0) or 0) + 1
        total_payments = _q2(total_payments + gross)
        total_tax = _q2(total_tax + vat_amount)

        records.append(
            {
                "sale_id": int(getattr(sale, "id", 0) or 0),
                "sold_at": sale.sold_at.isoformat() if getattr(sale, "sold_at", None) else None,
                "customer_name": str(getattr(sale, "customer_name", "") or "").strip() or None,
                "payment_method": method,
                "subtotal": float(getattr(sale, "subtotal", 0) or 0),
                "discount": float(getattr(sale, "discount", 0) or 0),
                "tax": float(getattr(sale, "tax", 0) or 0),
                "total": float(gross),
            }
        )

    method_order = ["CASH", "CARD", "EFT", "ONLINE", "ACCOUNT"]
    ordered_methods = [m for m in method_order if m in method_totals] + [m for m in sorted(method_totals.keys()) if m not in method_order]
    payment_methods = [
        {
            "method": method,
            "count": int(method_counts.get(method, 0) or 0),
            "total": float(method_totals.get(method, Decimal("0.00"))),
        }
        for method in ordered_methods
    ]

    registry = _get_pro_shop_exports_registry(db)
    export_info = registry.get(target_date.isoformat()) if isinstance(registry, dict) else None
    already_exported = isinstance(export_info, dict)

    return {
        "date": target_date.strftime("%Y-%m-%d"),
        "total_payments": float(total_payments),
        "total_tax": float(total_tax),
        "transaction_count": len(records),
        "payment_methods": payment_methods,
        "records": records,
        "already_exported": bool(already_exported),
        "export_batch_ref": str(export_info.get("batch_ref") or "").strip() if already_exported else None,
        "exported_at": str(export_info.get("exported_at") or "").strip() if already_exported else None,
    }


@router.get("/export-csv-pro-shop")
def export_pro_shop_payments_csv(
    request: Request,
    export_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    force: int = Query(0, description="1 to allow re-export for the same payment date (not recommended)"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    target_date = _parse_ymd_date(export_date, field_name="export_date")
    club_id = int(getattr(db, "info", {}).get("club_id") or 0)
    if club_id <= 0:
        raise HTTPException(status_code=400, detail="club_id is required")

    settings = get_accounting_settings(db)
    layout = _require_pastel_layout(db)
    mappings = _require_pastel_mappings(db)

    date_key = target_date.isoformat()
    exports_registry = _get_pro_shop_exports_registry(db)
    previous_export = exports_registry.get(date_key) if isinstance(exports_registry, dict) else None
    if not force and isinstance(previous_export, dict):
        prev_batch = str(previous_export.get("batch_ref") or "").strip()
        detail = "This pro shop payment date already looks exported. Use force=1 to re-export."
        if prev_batch:
            detail = f"This pro shop payment date already looks exported ({prev_batch}). Use force=1 to re-export."
        raise HTTPException(status_code=409, detail=detail)

    try:
        sales = (
            db.query(ProShopSale)
            .filter(ProShopSale.club_id == club_id, func.date(ProShopSale.sold_at) == target_date)
            .order_by(ProShopSale.sold_at.asc(), ProShopSale.id.asc())
            .all()
        )
    except Exception:
        raise HTTPException(status_code=503, detail="Database connection unavailable")

    if not sales:
        raise HTTPException(status_code=404, detail=f"No pro shop sales found for {target_date}")

    try:
        vat_rate = Decimal(str(settings.vat_rate if getattr(settings, "vat_rate", None) is not None else 0.15))
    except Exception:
        vat_rate = Decimal("0.15")
    if vat_rate < 0:
        vat_rate = Decimal("0")

    gross_by_method: dict[str, Decimal] = {}
    method_counts: dict[str, int] = {}
    sale_ids: list[int] = []
    gross_total = Decimal("0.00")
    net_total = Decimal("0.00")
    vat_total = Decimal("0.00")

    for sale in sales:
        gross = _q2(_to_decimal(getattr(sale, "total", 0) or 0))
        if gross <= 0:
            continue
        method = _normalize_payment_method_for_journal(getattr(sale, "payment_method", None))
        explicit_tax = _q2(_to_decimal(getattr(sale, "tax", 0) or 0))
        net_amount, vat_amount = _pro_shop_tax_split(gross, explicit_tax, vat_rate)

        gross_by_method[method] = _q2(gross_by_method.get(method, Decimal("0.00")) + gross)
        method_counts[method] = int(method_counts.get(method, 0) or 0) + 1
        gross_total = _q2(gross_total + gross)
        net_total = _q2(net_total + net_amount)
        vat_total = _q2(vat_total + vat_amount)
        if getattr(sale, "id", None):
            sale_ids.append(int(sale.id))

    if not gross_by_method:
        raise HTTPException(status_code=404, detail=f"No positive-value pro shop payments found for {target_date}")

    if _q2(net_total + vat_total) != gross_total:
        raise HTTPException(status_code=500, detail="Pro shop VAT split does not balance to gross total after rounding.")

    debit_gl = mappings.get("debit_gl") or {}
    used_methods = sorted(gross_by_method.keys())
    missing_debits = [method for method in used_methods if not str(debit_gl.get(method, "") or "").strip()]
    if missing_debits:
        raise HTTPException(
            status_code=400,
            detail=f"Missing debit GL mapping for pro shop payment method(s): {', '.join(missing_debits)}",
        )

    revenue_by_fee_type = (mappings.get("revenue_gl") or {}) if isinstance(mappings.get("revenue_gl"), dict) else {}
    revenue_gl = str(
        revenue_by_fee_type.get("pro_shop")
        or revenue_by_fee_type.get("retail")
        or getattr(settings, "green_fees_gl", None)
        or ""
    ).strip()
    if not revenue_gl:
        raise HTTPException(
            status_code=400,
            detail="Missing revenue GL mapping for pro_shop. Set revenue_gl.pro_shop in Pastel mappings or configure Green fees GL.",
        )

    vat_output_gl = str((mappings.get("vat_output_gl") or "")).strip()
    if not vat_output_gl:
        raise HTTPException(status_code=400, detail="Missing VAT output GL account in Pastel mappings.")

    batch_ref = f"GREENLINK_PROSHOP_{target_date.strftime('%Y%m%d')}"
    batch_desc = _clean_text(f"Pro shop takings {target_date.strftime('%Y-%m-%d')}", max_len=60)

    lines: list[dict] = []
    method_order = ["CASH", "CARD", "EFT", "ONLINE", "ACCOUNT"]
    ordered_methods = [m for m in method_order if m in gross_by_method] + [m for m in used_methods if m not in method_order]
    for method in ordered_methods:
        lines.append(
            {
                "account": str(debit_gl.get(method) or "").strip(),
                "debit": gross_by_method[method],
                "credit": Decimal("0.00"),
                "ref": _clean_text(method, max_len=20),
                "desc": _clean_text(f"{batch_desc} {method}", max_len=60),
            }
        )

    lines.append(
        {
            "account": revenue_gl,
            "debit": Decimal("0.00"),
            "credit": net_total,
            "ref": "PRO SHOP",
            "desc": _clean_text(f"{batch_desc} SALES", max_len=60),
            "vat": vat_total,
        }
    )

    if vat_total != 0:
        lines.append(
            {
                "account": vat_output_gl,
                "debit": Decimal("0.00"),
                "credit": vat_total,
                "ref": "VAT CONT",
                "desc": _clean_text(f"Output VAT {target_date.strftime('%Y-%m-%d')}", max_len=60),
            }
        )

    debit_sum = _q2(sum((line["debit"] for line in lines), Decimal("0.00")))
    credit_sum = _q2(sum((line["credit"] for line in lines), Decimal("0.00")))
    if debit_sum != credit_sum:
        raise HTTPException(status_code=500, detail="Pro shop journal is out of balance after rounding.")

    csv_text = _render_pastel_journal_csv(
        layout=layout,
        mappings=mappings,
        lines=lines,
        target_date=target_date,
        batch_ref=batch_ref,
        vat_rate=vat_rate,
    )

    try:
        exports_registry[date_key] = {
            "batch_ref": batch_ref,
            "exported_at": datetime.utcnow().isoformat(),
            "sales_count": len(set(sale_ids)),
            "gross_total": float(gross_total),
            "net_total": float(net_total),
            "vat_total": float(vat_total),
        }
        _set_pro_shop_exports_registry(db, exports_registry)

        _audit_cashbook_event(
            db,
            request,
            admin,
            action="cashbook.pro_shop_csv_exported",
            entity_type="cashbook_export",
            entity_id=batch_ref,
            payload={
                "date": target_date.isoformat(),
                "batch_ref": batch_ref,
                "force": int(force or 0),
                "sales_count": len(set(sale_ids)),
                "method_counts": {k: int(v) for k, v in method_counts.items()},
                "gross_total": float(gross_total),
                "net_total": float(net_total),
                "vat_total": float(vat_total),
            },
        )
        db.commit()
    except Exception as e:
        log_event(
            "warning",
            "cashbook.pro_shop_export.mark_failed",
            request_id=_request_id(request),
            error_type=type(e).__name__,
            error=str(e)[:240],
            batch_ref=batch_ref,
        )

    file_name = f"PASTEL_JOURNAL_PROSHOP_GREENLINK_{target_date.strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([csv_text.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={file_name}",
            "X-GreenLink-BatchRef": batch_ref,
        },
    )


@router.get("/export-excel")
def export_daily_payments_excel(
    export_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    """Export daily payments to Excel file"""
    
    # Parse date
    if export_date:
        try:
            target_date = datetime.strptime(export_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()
    
    # Get bookings
    try:
        bookings = get_active_completed_bookings(db, target_date)
    except Exception as e:
        raise HTTPException(status_code=503, detail="Database connection unavailable")
    
    if not bookings:
        raise HTTPException(status_code=404, detail=f"No payments found for {target_date}")
    
    # Convert to payment records
    settings = get_accounting_settings(db)
    records = [create_payment_record(booking, settings, payment_date=target_date) for booking in bookings]
    
    # Create Excel workbook
    excel_file = create_excel_workbook(records, target_date.strftime("%d/%m/%Y"))
    
    # Generate filename
    filename = f"Cashbook_Payments_{target_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        iter([excel_file.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export-csv")
def export_daily_payments_csv(
    request: Request,
    export_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    force: int = Query(0, description="1 to allow re-export for the same payment date (not recommended)"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    """
    Export a balanced daily VAT journal in the club's Pastel Partner journal import layout.

    Accounting rule: transaction date is the payment date (ledger_entries.created_at), not the tee time date.
    """
    if export_date:
        try:
            target_date = datetime.strptime(export_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()

    settings = get_accounting_settings(db)
    layout = _require_pastel_layout(db)
    mappings = _require_pastel_mappings(db)

    # Idempotency: avoid double-posting the same day (unless forced).
    if not force:
        exported_count = (
            db.query(func.count(LedgerEntry.id))
            .filter(func.date(LedgerEntry.created_at) == target_date, LedgerEntry.pastel_synced == 1)
            .scalar()
            or 0
        )
        if exported_count:
            raise HTTPException(status_code=409, detail="This payment date already looks exported. Reopen the day or use force=1.")

    try:
        rows = (
            db.query(LedgerEntry, Booking, FeeCategory, LedgerEntryMeta)
            .join(Booking, LedgerEntry.booking_id == Booking.id)
            .outerjoin(FeeCategory, FeeCategory.id == Booking.fee_category_id)
            .outerjoin(LedgerEntryMeta, LedgerEntryMeta.ledger_entry_id == LedgerEntry.id)
            .filter(
                LedgerEntry.booking_id.isnot(None),
                func.date(LedgerEntry.created_at) == target_date,
                Booking.status.in_([BookingStatus.checked_in, BookingStatus.completed]),
            )
            .all()
        )
    except Exception:
        raise HTTPException(status_code=503, detail="Database connection unavailable")

    if not rows:
        raise HTTPException(status_code=404, detail=f"No payments found for {target_date}")

    journal = build_daily_journal_lines(
        rows=rows,
        settings=settings,
        mappings=mappings,
        target_date=target_date,
        clean_text=_clean_text,
    )
    vat_output_gl = mappings.get("vat_output_gl")
    lines = journal.lines
    gross_total = journal.gross_total
    net_total = journal.net_total
    vat_total = journal.vat_total
    batch_ref = journal.batch_ref
    rate = journal.rate
    ledger_entry_ids = journal.ledger_entry_ids
    booking_ids = journal.booking_ids

    # Render CSV according to stored Pastel layout.
    columns = layout.get("columns") or []
    column_map = layout.get("column_map") or {}
    header_to_idx = {str(h): i for i, h in enumerate(columns)}

    has_debit = bool(column_map.get("debit")) and column_map.get("debit") in header_to_idx
    has_credit = bool(column_map.get("credit")) and column_map.get("credit") in header_to_idx
    has_amount = bool(column_map.get("amount")) and column_map.get("amount") in header_to_idx

    required = ["date", "account"]
    for key in required:
        header = column_map.get(key)
        if not header or header not in header_to_idx:
            raise HTTPException(status_code=500, detail=f"Pastel layout missing required column for '{key}'. Re-upload layout CSV.")

    # Layouts typically have either (Debit,Credit) columns or a single signed Amount column.
    if not ((has_debit and has_credit) or has_amount):
        raise HTTPException(
            status_code=500,
            detail="Pastel layout must include Debit/Credit columns or an Amount column. Re-upload a Batch->Export sample that matches your import layout.",
        )

    date_format = layout.get("date_format")
    date_value = _format_date_for_layout(target_date, date_format)
    delimiter = layout.get("delimiter") or ","
    has_header = bool(layout.get("has_header", True))
    template_row = layout.get("template_row")
    has_template = isinstance(template_row, list) and len(template_row) == len(columns)
    sample_rows = layout.get("sample_rows") or []
    has_samples = isinstance(sample_rows, list) and bool(sample_rows)

    # Pastel "Batch -> Export" templates are great for column order, but the first data row can
    # contain row-specific values (e.g. project/tax codes) that should NOT be cloned onto every
    # exported row. To avoid import errors, only keep values that appear constant across the
    # sampled rows; everything else starts blank unless we explicitly set it via column_map.
    constant_template: list[str | None] = [None for _ in columns]
    if has_template and has_samples:
        try:
            # Only consider rows that match the column count.
            rows_for_constants = [r for r in sample_rows if isinstance(r, list) and len(r) == len(columns)]
            if rows_for_constants:
                for i in range(len(columns)):
                    vals = [str(r[i] or "").strip() for r in rows_for_constants]
                    if not vals:
                        continue
                    first = vals[0]
                    if first and all(v == first for v in vals[1:]):
                        # Preserve original formatting from template_row (not stripped).
                        constant_template[i] = str(template_row[i] or "")
        except Exception:
            constant_template = [None for _ in columns]

    def _base_row() -> list:
        if has_template and any(v is not None for v in constant_template):
            row = ["" for _ in columns]
            for i, v in enumerate(constant_template):
                if v is not None and i < len(row):
                    row[i] = v
            return row
        if has_template:
            return [str(x or "") for x in list(template_row)]
        return ["" for _ in columns]

    def _set(row: list, key: str, value: str) -> None:
        header = column_map.get(key)
        if not header:
            return
        idx = header_to_idx.get(header)
        if idx is None:
            return
        row[idx] = value

    # Some Pastel layouts repeat the signed amount in multiple columns.
    amount_mirror_headers: list[str] = []
    try:
        amount_mirror_headers = list(((layout.get("inferred") or {}).get("amount_mirrors") or []))
    except Exception:
        amount_mirror_headers = []

    def _set_amount_mirrors(row: list, value: str) -> None:
        for h in amount_mirror_headers:
            idx = header_to_idx.get(str(h))
            if idx is None:
                continue
            if idx < len(row):
                row[idx] = value

    out = StringIO(newline="")
    writer = csv.writer(out, delimiter=delimiter, lineterminator="\r\n")
    if has_header:
        writer.writerow(columns)

    tax_type_sales = str((mappings.get("tax_type") or "")).strip()
    tax_type_sales_is_numeric = bool(tax_type_sales) and tax_type_sales.isdigit()
    amount_sign = str((mappings.get("amount_sign") or "")).strip().lower() or "debit_positive"
    vat_output_gl_fmt = _format_gl_for_layout(str(vat_output_gl or "").strip(), layout) if vat_output_gl else ""
    for line in lines:
        row = _base_row()
        _set(row, "date", date_value)
        # Reference/Description
        if column_map.get("reference") and column_map.get("reference") in header_to_idx:
            _set(row, "reference", batch_ref if has_header else str(line.get("ref") or batch_ref))
        if column_map.get("description") and column_map.get("description") in header_to_idx:
            # If the template already has a short constant description (e.g. "Payment"), keep it.
            if has_template:
                d_idx = header_to_idx[column_map.get("description")]
                existing = str(row[d_idx] or "").strip() if d_idx < len(row) else ""
                if existing and len(existing) <= 20:
                    pass
                else:
                    _set(row, "description", str(line["desc"]))
            else:
                _set(row, "description", str(line["desc"]))

        account_value = _format_gl_for_layout(str(line["account"]), layout)
        _set(row, "account", account_value)

        # Determine whether this is a revenue line (taxable) vs debit line or VAT control line.
        is_debit_line = bool(line["debit"] and _q2(line["debit"]) != 0)
        is_credit_line = bool((not is_debit_line) and line["credit"] and _q2(line["credit"]) != 0)
        is_vat_control_line = bool(vat_output_gl_fmt) and account_value.strip() == vat_output_gl_fmt
        is_revenue_line = bool(is_credit_line and not is_vat_control_line)
        apply_sales_tax = bool(is_revenue_line and tax_type_sales)

        # Amount (or Debit/Credit)
        amt = None
        if has_amount:
            # Signed-amount layout.
            amt = Decimal("0.00")
            if is_debit_line:
                amt = _q2(line["debit"])
                if amount_sign == "debit_negative":
                    amt = -amt
            else:
                amt = _q2(line["credit"])
                # Credit is opposite sign to debit.
                if amount_sign == "debit_positive":
                    amt = -amt
                else:
                    amt = +amt
            amt_str = _money_str(amt)
            _set(row, "amount", amt_str)
            _set_amount_mirrors(row, amt_str)
            if has_debit:
                _set(row, "debit", "")
            if has_credit:
                _set(row, "credit", "")
        else:
            # Debit/Credit layout.
            if is_debit_line:
                _set(row, "debit", _money_str(line["debit"]))
                _set(row, "credit", "")
            else:
                _set(row, "debit", "")
                _set(row, "credit", _money_str(line["credit"]))

        # Tax fields (layout-specific).
        # Many Pastel layouts expect explicit "00"/"01" + numeric zeroes, not blanks.
        if column_map.get("tax_type") and column_map.get("tax_type") in header_to_idx:
            if apply_sales_tax and tax_type_sales:
                _set(row, "tax_type", tax_type_sales)
            else:
                # If the configured sales tax "type" looks numeric (e.g. "01"), non-tax lines
                # typically use "00". If it looks like an alphanumeric code (e.g. "GOV01"),
                # keep non-tax lines blank to avoid triggering validation on unrelated fields
                # in some Pastel layouts.
                _set(row, "tax_type", "00" if tax_type_sales_is_numeric else "")

        if column_map.get("tax_flag") and column_map.get("tax_flag") in header_to_idx:
            _set(row, "tax_flag", "1" if apply_sales_tax else "0")

        if column_map.get("tax_amount") and column_map.get("tax_amount") in header_to_idx:
            if apply_sales_tax:
                # Prefer per-line VAT (aggregated with transaction-level rounding), else fall back.
                vat_amt = _q2(_to_decimal(line.get("vat", 0) or 0))
                if vat_amt == 0:
                    base_net = _q2(line["credit"]) if is_credit_line else _q2(line["debit"])
                    vat_amt = _q2(base_net * rate) if rate > 0 else Decimal("0.00")
                # Match sign to the signed amount when available.
                if amt is not None and amt < 0:
                    vat_amt = -vat_amt
                _set(row, "tax_amount", _money_str_zero(vat_amt))
            else:
                _set(row, "tax_amount", "0")

        writer.writerow(row)

    csv_text = out.getvalue()

    # Mark ledger entries as exported (idempotency + audit trail).
    try:
        if ledger_entry_ids:
            db.query(LedgerEntry).filter(LedgerEntry.id.in_(list(set(ledger_entry_ids)))).update(
                {
                    LedgerEntry.pastel_synced: 1,
                    LedgerEntry.pastel_transaction_id: batch_ref,
                },
                synchronize_session=False,
            )
            db.commit()
            _audit_cashbook_event(
                db,
                request,
                admin,
                action="cashbook.csv_exported",
                entity_type="cashbook_export",
                entity_id=batch_ref,
                payload={
                    "date": target_date.isoformat(),
                    "batch_ref": batch_ref,
                    "force": int(force or 0),
                    "ledger_entries": len(set(ledger_entry_ids)),
                    "bookings": len(set(booking_ids)),
                    "gross_total": float(gross_total),
                    "net_total": float(net_total),
                    "vat_total": float(vat_total),
                },
            )
            db.commit()
    except Exception as e:
        log_event(
            "warning",
            "cashbook.export.mark_failed",
            request_id=_request_id(request),
            error_type=type(e).__name__,
            error=str(e)[:240],
            batch_ref=batch_ref,
        )

    file_name = f"PASTEL_JOURNAL_GREENLINK_{target_date.strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([csv_text.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={file_name}",
            "X-GreenLink-BatchRef": batch_ref,
        },
    )


@router.get("/export-preview")
def export_daily_payments_preview(
    export_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    admin: User = Depends(verify_admin),
    db: Session = Depends(get_db),
):
    """
    Preview the daily Pastel journal (no file write, no DB flags updated).

    Uses the same logic as /export-csv so the UI can show a "what will import" preview.
    """
    if export_date:
        try:
            target_date = datetime.strptime(export_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()

    club_id = int(getattr(db, "info", {}).get("club_id") or 0)
    settings = get_accounting_settings(db)
    mapping_status = get_export_mapping_status(db, club_id=club_id) if club_id else {"configured": False, "layout_configured": False, "mappings_configured": False}
    finance_semantics = build_finance_semantics_metadata(mapping_status)
    layout = _require_pastel_layout(db)
    mappings = _require_pastel_mappings(db)

    try:
        rows = (
            db.query(LedgerEntry, Booking, FeeCategory, LedgerEntryMeta)
            .join(Booking, LedgerEntry.booking_id == Booking.id)
            .outerjoin(FeeCategory, FeeCategory.id == Booking.fee_category_id)
            .outerjoin(LedgerEntryMeta, LedgerEntryMeta.ledger_entry_id == LedgerEntry.id)
            .filter(
                LedgerEntry.booking_id.isnot(None),
                func.date(LedgerEntry.created_at) == target_date,
                Booking.status.in_([BookingStatus.checked_in, BookingStatus.completed]),
            )
            .all()
        )
    except Exception:
        raise HTTPException(status_code=503, detail="Database connection unavailable")

    if not rows:
        raise HTTPException(status_code=404, detail=f"No payments found for {target_date}")

    preview_finance_states = [
        build_ledger_entry_finance_state(
            pastel_synced=bool(getattr(ledger_entry, "pastel_synced", False)),
            payment_method=str(getattr(meta, "payment_method", "") or "").strip().upper() or None,
            mapping_status=mapping_status,
        )
        for ledger_entry, _booking, _fee_cat, meta in rows
    ]

    journal = build_daily_journal_lines(
        rows=rows,
        settings=settings,
        mappings=mappings,
        target_date=target_date,
        clean_text=_clean_text,
    )

    vat_output_gl = mappings.get("vat_output_gl")
    if not str(vat_output_gl or "").strip():
        raise HTTPException(status_code=400, detail="Missing VAT output GL account in Pastel mappings.")

    lines = journal.lines
    gross_total = journal.gross_total
    net_total = journal.net_total
    vat_total = journal.vat_total
    batch_ref = journal.batch_ref
    rate = journal.rate

    # Render CSV rows for an exact preview (matches file output).
    columns = layout.get("columns") or []
    column_map = layout.get("column_map") or {}
    header_to_idx = {str(h): i for i, h in enumerate(columns)}

    has_amount = bool(column_map.get("amount")) and column_map.get("amount") in header_to_idx
    has_debit = bool(column_map.get("debit")) and column_map.get("debit") in header_to_idx
    has_credit = bool(column_map.get("credit")) and column_map.get("credit") in header_to_idx

    date_format = layout.get("date_format")
    date_value = _format_date_for_layout(target_date, date_format)
    delimiter = layout.get("delimiter") or ","
    has_header = bool(layout.get("has_header", True))
    template_row = layout.get("template_row")
    has_template = isinstance(template_row, list) and len(template_row) == len(columns)
    sample_rows = layout.get("sample_rows") or []
    has_samples = isinstance(sample_rows, list) and bool(sample_rows)

    constant_template: list[str | None] = [None for _ in columns]
    if has_template and has_samples:
        try:
            rows_for_constants = [r for r in sample_rows if isinstance(r, list) and len(r) == len(columns)]
            if rows_for_constants:
                for i in range(len(columns)):
                    vals = [str(r[i] or "").strip() for r in rows_for_constants]
                    if not vals:
                        continue
                    first = vals[0]
                    if first and all(v == first for v in vals[1:]):
                        constant_template[i] = str(template_row[i] or "")
        except Exception:
            constant_template = [None for _ in columns]

    def _base_row() -> list:
        if has_template and any(v is not None for v in constant_template):
            row = ["" for _ in columns]
            for i, v in enumerate(constant_template):
                if v is not None and i < len(row):
                    row[i] = v
            return row
        if has_template:
            return [str(x or "") for x in list(template_row)]
        return ["" for _ in columns]

    def _set(row: list, key: str, value: str) -> None:
        header = column_map.get(key)
        if not header:
            return
        idx = header_to_idx.get(header)
        if idx is None:
            return
        row[idx] = value

    amount_mirror_headers: list[str] = []
    try:
        amount_mirror_headers = list(((layout.get("inferred") or {}).get("amount_mirrors") or []))
    except Exception:
        amount_mirror_headers = []

    def _set_amount_mirrors(row: list, value: str) -> None:
        for h in amount_mirror_headers:
            idx = header_to_idx.get(str(h))
            if idx is None:
                continue
            if idx < len(row):
                row[idx] = value

    out = StringIO(newline="")
    writer = csv.writer(out, delimiter=delimiter, lineterminator="\r\n")
    rendered_rows: list[list[str]] = []
    if has_header:
        writer.writerow(columns)
        rendered_rows.append([str(c or "") for c in columns])

    tax_type_sales = str((mappings.get("tax_type") or "")).strip()
    tax_type_sales_is_numeric = bool(tax_type_sales) and tax_type_sales.isdigit()
    amount_sign = str((mappings.get("amount_sign") or "")).strip().lower() or "debit_positive"
    vat_output_gl_fmt = _format_gl_for_layout(str(vat_output_gl or "").strip(), layout) if vat_output_gl else ""

    for line in lines:
        row = _base_row()
        _set(row, "date", date_value)
        if column_map.get("reference") and column_map.get("reference") in header_to_idx:
            _set(row, "reference", batch_ref if has_header else str(line.get("ref") or batch_ref))
        if column_map.get("description") and column_map.get("description") in header_to_idx:
            if has_template:
                d_idx = header_to_idx[column_map.get("description")]
                existing = str(row[d_idx] or "").strip() if d_idx < len(row) else ""
                if existing and len(existing) <= 20:
                    pass
                else:
                    _set(row, "description", str(line["desc"]))
            else:
                _set(row, "description", str(line["desc"]))

        account_value = _format_gl_for_layout(str(line["account"]), layout)
        _set(row, "account", account_value)

        is_debit_line = bool(line["debit"] and _q2(line["debit"]) != 0)
        is_credit_line = bool((not is_debit_line) and line["credit"] and _q2(line["credit"]) != 0)
        is_vat_control_line = bool(vat_output_gl_fmt) and account_value.strip() == vat_output_gl_fmt
        is_revenue_line = bool(is_credit_line and not is_vat_control_line)
        apply_sales_tax = bool(is_revenue_line and tax_type_sales)

        amt = None
        if has_amount:
            amt = Decimal("0.00")
            if is_debit_line:
                amt = _q2(line["debit"])
                if amount_sign == "debit_negative":
                    amt = -amt
            else:
                amt = _q2(line["credit"])
                if amount_sign == "debit_positive":
                    amt = -amt
                else:
                    amt = +amt
            amt_str = _money_str(amt)
            _set(row, "amount", amt_str)
            _set_amount_mirrors(row, amt_str)
            if has_debit:
                _set(row, "debit", "")
            if has_credit:
                _set(row, "credit", "")
        else:
            if is_debit_line:
                _set(row, "debit", _money_str(line["debit"]))
                _set(row, "credit", "")
            else:
                _set(row, "debit", "")
                _set(row, "credit", _money_str(line["credit"]))

        if column_map.get("tax_type") and column_map.get("tax_type") in header_to_idx:
            if apply_sales_tax and tax_type_sales:
                _set(row, "tax_type", tax_type_sales)
            else:
                _set(row, "tax_type", "00" if tax_type_sales_is_numeric else "")

        if column_map.get("tax_flag") and column_map.get("tax_flag") in header_to_idx:
            _set(row, "tax_flag", "1" if apply_sales_tax else "0")

        if column_map.get("tax_amount") and column_map.get("tax_amount") in header_to_idx:
            if apply_sales_tax:
                vat_amt = _q2(_to_decimal(line.get("vat", 0) or 0))
                if vat_amt == 0:
                    base_net = _q2(line["credit"]) if is_credit_line else _q2(line["debit"])
                    vat_amt = _q2(base_net * rate) if rate > 0 else Decimal("0.00")
                if amt is not None and amt < 0:
                    vat_amt = -vat_amt
                _set(row, "tax_amount", _money_str_zero(vat_amt))
            else:
                _set(row, "tax_amount", "0")

        writer.writerow(row)
        rendered_rows.append([str(x or "") for x in row])

    preview_lines: list[dict] = []
    for row in rendered_rows[1 if has_header else 0 :]:
        def _get(key: str) -> str:
            header = column_map.get(key)
            if not header:
                return ""
            idx = header_to_idx.get(header)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        preview_lines.append(
            {
                "date": _get("date"),
                "account": _get("account"),
                "reference": _get("reference"),
                "description": _get("description"),
                "amount": _get("amount"),
                "debit": _get("debit"),
                "credit": _get("credit"),
                "tax_type": _get("tax_type"),
                "tax_flag": _get("tax_flag"),
                "tax_amount": _get("tax_amount"),
            }
        )

    return {
        "date": target_date.strftime("%Y-%m-%d"),
        "batchRef": batch_ref,
        "finance_semantics": finance_semantics,
        "finance_state_summary": summarize_ledger_finance_states(preview_finance_states),
        "totals": {"gross": float(gross_total), "net": float(net_total), "vat": float(vat_total)},
        "journal_lines": preview_lines,
    }


@router.get("/close-status")
def get_close_status(
    summary_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    if summary_date:
        try:
            target_date = datetime.strptime(summary_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()

    club_id = int(getattr(db, "info", {}).get("club_id") or 0)
    mapping_status = get_export_mapping_status(db, club_id=club_id) if club_id else {"configured": False, "layout_configured": False, "mappings_configured": False}
    finance_semantics = build_finance_semantics_metadata(mapping_status)
    base_ledger_query = db.query(LedgerEntry).filter(
        LedgerEntry.club_id == club_id,
        LedgerEntry.booking_id.isnot(None),
        func.date(LedgerEntry.created_at) == target_date,
    )
    total_rows = int(base_ledger_query.count() or 0)
    exported_rows = int(base_ledger_query.filter(LedgerEntry.pastel_synced == 1).count() or 0)
    pending_export_rows = max(0, total_rows - exported_rows)
    missing_payment_method_rows = int(
        db.query(LedgerEntry)
        .outerjoin(LedgerEntryMeta, LedgerEntryMeta.ledger_entry_id == LedgerEntry.id)
        .filter(
            LedgerEntry.club_id == club_id,
            LedgerEntry.booking_id.isnot(None),
            func.date(LedgerEntry.created_at) == target_date,
            func.coalesce(LedgerEntry.pastel_synced, 0) != 1,
            func.length(func.trim(func.coalesce(LedgerEntryMeta.payment_method, ""))) == 0,
        )
        .count()
        or 0
    )
    mapping_configured = bool(mapping_status.get("configured"))
    export_ready_rows = max(0, pending_export_rows - missing_payment_method_rows) if mapping_configured else 0
    missing_mapping_rows = pending_export_rows if not mapping_configured else 0
    blocked_rows = max(0, pending_export_rows - export_ready_rows)
    finance_state_summary = {
        "total_rows": total_rows,
        "exported_rows": exported_rows,
        "pending_export_rows": pending_export_rows,
        "export_ready_rows": export_ready_rows,
        "blocked_rows": blocked_rows,
        "missing_payment_method_rows": missing_payment_method_rows,
        "missing_mapping_rows": missing_mapping_rows,
    }

    close = db.query(DayClose).filter(DayClose.close_date == target_date).order_by(DayClose.id.desc()).first()
    if not close:
        return {
            "date": target_date.strftime("%Y-%m-%d"),
            "is_closed": False,
            "finance_semantics": finance_semantics,
            "finance_state_summary": finance_state_summary,
        }

    return {
        "date": target_date.strftime("%Y-%m-%d"),
        "is_closed": close.status == "closed",
        "status": close.status,
        "closed_at": close.closed_at.isoformat() if close.closed_at else None,
        "closed_by_user_id": close.closed_by_user_id,
        "export_batch_id": close.export_batch_id,
        "export_filename": close.export_filename,
        "auto_push": bool(close.auto_push),
        "finance_semantics": finance_semantics,
        "finance_state_summary": finance_state_summary,
    }


@router.get("/settings")
def get_accounting_settings_api(
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin)
):
    settings = get_accounting_settings(db)
    gl_reference = _get_gl_account_reference(db)
    return {
        "green_fees_gl": settings.green_fees_gl,
        "cashbook_contra_gl": settings.cashbook_contra_gl,
        "vat_rate": settings.vat_rate,
        "tax_type": settings.tax_type,
        "cashbook_name": settings.cashbook_name,
        "updated_at": settings.updated_at.isoformat() if settings.updated_at else None,
        "gl_reference": gl_reference,
    }


@router.put("/settings")
def update_accounting_settings_api(
    payload: AccountingSettingsPayload,
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin)
):
    settings = get_accounting_settings(db)
    if payload.green_fees_gl is not None:
        settings.green_fees_gl = payload.green_fees_gl.strip()
    if payload.cashbook_contra_gl is not None:
        settings.cashbook_contra_gl = payload.cashbook_contra_gl.strip()
    if payload.vat_rate is not None:
        settings.vat_rate = payload.vat_rate
    if payload.tax_type is not None:
        settings.tax_type = payload.tax_type
    if payload.cashbook_name is not None:
        settings.cashbook_name = payload.cashbook_name.strip()
    settings.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(settings)
    return {
        "status": "success",
        "settings": {
            "green_fees_gl": settings.green_fees_gl,
            "cashbook_contra_gl": settings.cashbook_contra_gl,
            "vat_rate": settings.vat_rate,
            "tax_type": settings.tax_type,
            "cashbook_name": settings.cashbook_name
        }
    }


@router.post("/close-day")
def close_day(
    request: Request,
    close_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin)
):
    club_id = int(getattr(db, "info", {}).get("club_id") or 0)
    if club_id <= 0:
        raise HTTPException(status_code=400, detail="club_id is required")

    if close_date:
        try:
            target_date = datetime.strptime(close_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()

    existing = db.query(DayClose).filter(
        DayClose.close_date == target_date,
        DayClose.status == "closed"
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Day is already closed")

    bookings = get_active_completed_bookings(db, target_date)
    booking_ids = [b.id for b in bookings]
    batch_id = f"GL-{target_date.strftime('%Y%m%d')}-{datetime.utcnow().strftime('%H%M%S')}"

    # NOTE: Do not mark ledger entries as exported here.
    # /cashbook/export-csv writes the CSV and then marks entries as exported.

    filename = f"Cashbook_Payments_{target_date.strftime('%Y%m%d')}.csv"
    close = db.query(DayClose).filter(DayClose.close_date == target_date).first()
    if close:
        close.status = "closed"
        close.closed_by_user_id = admin.id
        close.closed_at = datetime.utcnow()
        close.export_method = "cashbook"
        close.export_batch_id = batch_id
        close.export_filename = filename
        close.auto_push = 0
    else:
        close = DayClose(
            club_id=club_id,
            close_date=target_date,
            status="closed",
            closed_by_user_id=admin.id,
            closed_at=datetime.utcnow(),
            export_method="cashbook",
            export_batch_id=batch_id,
            export_filename=filename,
            auto_push=0
        )
        db.add(close)

    _audit_cashbook_event(
        db,
        request,
        admin,
        action="cashbook.day_closed",
        entity_type="day_close",
        entity_id=target_date.isoformat(),
        payload={
            "date": target_date.isoformat(),
            "batch_id": batch_id,
            "bookings": len(booking_ids),
            "export_filename": filename,
        },
    )
    db.commit()

    return {
        "status": "closed",
        "date": target_date.strftime("%Y-%m-%d"),
        "batch_id": batch_id,
        "bookings": len(booking_ids),
        "export_filename": filename
    }


@router.post("/reopen-day")
def reopen_day(
    request: Request,
    reopen_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin)
):
    if reopen_date:
        try:
            target_date = datetime.strptime(reopen_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()

    close = db.query(DayClose).filter(
        DayClose.close_date == target_date,
        DayClose.status == "closed"
    ).first()
    if not close:
        raise HTTPException(status_code=404, detail="Day is not closed")

    bookings = get_active_completed_bookings(db, target_date)
    booking_ids = [b.id for b in bookings]
    if booking_ids:
        ledger_entries = db.query(LedgerEntry).filter(LedgerEntry.booking_id.in_(booking_ids)).all()
        for le in ledger_entries:
            le.pastel_synced = 0
            le.pastel_transaction_id = None

    exports_registry = _get_pro_shop_exports_registry(db)
    if isinstance(exports_registry, dict) and target_date.isoformat() in exports_registry:
        exports_registry.pop(target_date.isoformat(), None)
        _set_pro_shop_exports_registry(db, exports_registry)

    close.status = "reopened"
    close.reopened_by_user_id = admin.id
    close.reopened_at = datetime.utcnow()
    _audit_cashbook_event(
        db,
        request,
        admin,
        action="cashbook.day_reopened",
        entity_type="day_close",
        entity_id=target_date.isoformat(),
        payload={"date": target_date.isoformat(), "booking_count": len(booking_ids)},
    )
    db.commit()

    return {
        "status": "reopened",
        "date": target_date.strftime("%Y-%m-%d")
    }


@router.post("/finalize-day")
def finalize_day_payments(
    request: Request,
    finalize_date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format, defaults to today"),
    db: Session = Depends(get_db),
    admin: User = Depends(verify_admin),
):
    """
    Finalize payments for a day:
    1. Get all checked-in and completed bookings
    2. Create payment records
    3. Export to Excel
    4. Return file path and summary
    """
    
    # Parse date
    if finalize_date:
        try:
            target_date = datetime.strptime(finalize_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        target_date = date.today()
    
    # Get bookings
    try:
        bookings = get_active_completed_bookings(db, target_date)
    except Exception as e:
        return {
            "status": "error",
            "date": target_date.strftime("%Y-%m-%d"),
            "message": "Database connection unavailable",
            "transaction_count": 0,
            "total_amount": 0.0
        }
    
    if not bookings:
        return {
            "status": "no_data",
            "date": target_date.strftime("%Y-%m-%d"),
            "message": f"No payments found for {target_date}",
            "transaction_count": 0,
            "total_amount": 0.0
        }
    
    # Convert to payment records
    settings = get_accounting_settings(db)
    records = [create_payment_record(booking, settings, payment_date=target_date) for booking in bookings]
    
    # Calculate totals
    total_payments = sum(r.amount for r in records)
    total_tax = sum(r.tax_amount for r in records)
    _audit_cashbook_event(
        db,
        request,
        admin,
        action="cashbook.day_finalized",
        entity_type="cashbook_finalize",
        entity_id=target_date.isoformat(),
        payload={
            "date": target_date.isoformat(),
            "transaction_count": len(records),
            "total_amount": float(total_payments),
            "total_tax": float(total_tax),
        },
    )
    db.commit()
    log_event(
        "info",
        "cashbook.finalize_day",
        request_id=_request_id(request),
        actor_user_id=int(getattr(admin, "id", 0) or 0),
        date=target_date.isoformat(),
        transaction_count=len(records),
        total_amount=float(total_payments),
        total_tax=float(total_tax),
    )
    
    return {
        "status": "success",
        "date": target_date.strftime("%Y-%m-%d"),
        "transaction_count": len(records),
        "total_amount": total_payments,
        "total_tax": total_tax,
        "export_url": f"/cashbook/export-excel?export_date={target_date.strftime('%Y-%m-%d')}",
        "message": f"Successfully processed {len(records)} payments for {target_date.strftime('%Y-%m-%d')}"
    }
