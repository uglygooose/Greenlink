from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from app import models
from app.people import (
    classify_membership_group,
    member_identity_key,
    normalize_primary_operation,
    normalize_membership_status,
    parse_membership_date,
    parse_terms_days,
    parse_yes_no_flag,
)


DEFAULT_SETUP_DIRS = (
    "UMHLALI_OPERATING_INPUT_DIR",
    r"c:\Users\athom\OneDrive\Pictures\Greenlink\Umhlali Setup",
    r".\umhlali_setup",
)
GL_ACCOUNTS_FILE_TOKENS = ("glaccounts", "gl accounts", "chart of accounts")


@dataclass(frozen=True)
class SheetRef:
    path: Path
    sheet_name: str | None = None


@dataclass(frozen=True)
class SetupFiles:
    setup_dir: Path
    member_list: SheetRef | None
    account_customers: SheetRef | None
    golf_day_bookings: SheetRef | None
    staff_roles: SheetRef | None


def _clean_text(value: Any, *, max_len: int = 255) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if len(text) > max_len:
        text = text[:max_len]
    return text


def _to_float(value: Any) -> float | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.replace("R", "").replace(",", "").strip()
    m = re.search(r"[-+]?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _to_date(value: Any) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except Exception:
        return None


def _sheet_ref(path: Path | None, sheet_name: str | None = None) -> SheetRef | None:
    if path is None:
        return None
    return SheetRef(path=path, sheet_name=sheet_name)


def _workbook_has_sheets(path: Path, required: set[str]) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = {str(name or "").strip().lower() for name in wb.sheetnames}
        return required.issubset(names)
    except Exception:
        return False


def _candidate_workbook_paths() -> list[Path]:
    candidates: list[Path] = []

    import os

    env_value = _clean_text(os.getenv("UMHLALI_OPERATING_INPUT_DIR"))
    if env_value:
        candidates.append(Path(env_value))
    env_file = _clean_text(os.getenv("UMHLALI_OPERATING_INPUT_FILE"))
    if env_file:
        candidates.append(Path(env_file))
    for item in DEFAULT_SETUP_DIRS[1:]:
        candidates.append(Path(item))
    candidates.append(Path.home() / "Downloads")

    workbook_paths: list[Path] = []
    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file() and candidate.suffix.lower() == ".xlsx":
                workbook_paths.append(candidate)
            elif candidate.exists() and candidate.is_dir():
                workbook_paths.extend(sorted(candidate.glob("*.xlsx")))
        except Exception:
            continue

    return workbook_paths


def _find_setup_files() -> SetupFiles | None:
    workbook_paths = _candidate_workbook_paths()
    if not workbook_paths:
        return None

    required_clean_sheets = {
        "members_clean",
        "account_customers_clean",
        "golf_day_bookings_clean",
        "staff_users_clean",
    }
    for path in workbook_paths:
        if _workbook_has_sheets(path, required_clean_sheets):
            return SetupFiles(
                setup_dir=path.parent,
                member_list=_sheet_ref(path, "Members_Clean"),
                account_customers=_sheet_ref(path, "Account_Customers_Clean"),
                golf_day_bookings=_sheet_ref(path, "Golf_Day_Bookings_Clean"),
                staff_roles=_sheet_ref(path, "Staff_Users_Clean"),
            )

    setup_dir = workbook_paths[0].parent
    file_map = {p.name.lower(): p for p in workbook_paths}
    member = None
    account = None
    golf_day = None
    staff = None
    for lower_name, path in file_map.items():
        if "member list" in lower_name and "golf scape" in lower_name:
            member = path
        elif "account customers" in lower_name:
            account = path
        elif "golf day bookings" in lower_name:
            golf_day = path
        elif "staff user list" in lower_name:
            staff = path

    return SetupFiles(
        setup_dir=setup_dir,
        member_list=_sheet_ref(member),
        account_customers=_sheet_ref(account),
        golf_day_bookings=_sheet_ref(golf_day),
        staff_roles=_sheet_ref(staff),
    )


def find_umhlali_setup_files() -> SetupFiles | None:
    return _find_setup_files()


def find_umhlali_gl_accounts_file() -> Path | None:
    for path in _candidate_workbook_paths():
        lower_name = str(path.name or "").strip().lower()
        if any(token in lower_name for token in GL_ACCOUNTS_FILE_TOKENS):
            return path
    return None


def extract_gl_accounts_reference(path: Path) -> list[dict[str, str]]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []

    if not wb.sheetnames:
        return []

    ws = wb[wb.sheetnames[0]]
    out: list[dict[str, str]] = []
    seen: set[str] = set()

    for row in ws.iter_rows(values_only=True):
        account = _clean_text(row[1] if len(row) > 1 else None, max_len=40)
        description = _clean_text(row[2] if len(row) > 2 else None, max_len=200)
        if not account or not description:
            continue
        key = account.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append({"account": account, "description": description})
    return out


def _sheet_rows(source: SheetRef) -> list[list[Any]]:
    wb = openpyxl.load_workbook(source.path, data_only=True)
    preferred = str(source.sheet_name or "").strip()
    if preferred and preferred in wb.sheetnames:
        ws = wb[preferred]
    else:
        ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _header_index_map(header: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, v in enumerate(header):
        key = str(v or "").strip().lower()
        if not key:
            continue
        out[key] = i
    return out


def _extract_cells(row: list[Any], header_map: dict[str, int], *keys: str) -> Any:
    for key in keys:
        idx = header_map.get(str(key).strip().lower())
        if idx is None:
            continue
        if idx >= len(row):
            continue
        value = row[idx]
        if str(value or "").strip() == "":
            continue
        return value
    return None


def _find_header_row(rows: list[list[Any]], *required_keys: str) -> tuple[int, dict[str, int]] | None:
    required = {str(key or "").strip().lower() for key in required_keys if str(key or "").strip()}
    if not required:
        return None
    for idx, row in enumerate(rows[:20]):
        cells = {str(v or "").strip().lower() for v in row if str(v or "").strip()}
        if required.issubset(cells):
            return idx, _header_index_map(row)
    return None


def _find_members_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    return (
        _find_header_row(rows, "source_row_number", "first_name", "last_name", "membership_category_raw")
        or _find_header_row(rows, "first name", "last name", "membership")
    )


def _find_account_customers_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    return _find_header_row(rows, "customer_name", "account_code") or _find_header_row(rows, "name")


def _upsert_import_batch(db: Session, *, club_id: int, source: str, file_name: str | None) -> models.ImportBatch:
    batch = models.ImportBatch(
        club_id=int(club_id),
        kind="operational",
        source=source,
        file_name=file_name,
        imported_at=datetime.utcnow(),
    )
    db.add(batch)
    db.flush()
    return batch


def _upsert_account_customer(
    db: Session,
    *,
    club_id: int,
    name: str,
    account_code: str | None,
    billing_contact: str | None,
    terms_label: str | None,
    customer_type: str | None = None,
    operation_area: str | None = None,
    source_file: str | None = None,
    import_reference: str | None = None,
) -> tuple[models.AccountCustomer, bool]:
    row = None
    if account_code:
        row = (
            db.query(models.AccountCustomer)
            .filter(
                models.AccountCustomer.club_id == int(club_id),
                func.lower(models.AccountCustomer.account_code) == account_code.lower(),
            )
            .first()
        )
    if row is None:
        row = (
            db.query(models.AccountCustomer)
            .filter(
                models.AccountCustomer.club_id == int(club_id),
                func.lower(models.AccountCustomer.name) == name.lower(),
            )
            .first()
        )

    created = False
    if row is None:
        row = models.AccountCustomer(
            club_id=int(club_id),
            name=name,
            account_code=account_code,
            billing_contact=billing_contact,
            terms_label=terms_label,
            terms_days=parse_terms_days(terms_label),
            customer_type=customer_type,
            operation_area=operation_area,
            source_file=source_file,
            import_reference=import_reference,
            active=1,
        )
        db.add(row)
        db.flush()
        created = True
    else:
        row.name = name
        row.account_code = account_code
        row.billing_contact = billing_contact
        row.terms_label = terms_label
        row.terms_days = parse_terms_days(terms_label)
        row.customer_type = customer_type
        row.operation_area = operation_area
        row.source_file = source_file
        row.import_reference = import_reference
        row.updated_at = datetime.utcnow()
    return row, created


def _ingest_account_customers(
    db: Session,
    *,
    club_id: int,
    path: SheetRef,
) -> dict[str, int]:
    rows = _sheet_rows(path)
    if not rows:
        return {"rows": 0, "inserted": 0, "updated": 0}
    header_info = _find_account_customers_header(rows)
    if header_info is None:
        return {"rows": 0, "inserted": 0, "updated": 0}
    header_idx, header = header_info
    inserted = 0
    updated = 0
    touched_refs: set[str] = set()
    for row in rows[header_idx + 1 :]:
        name = _clean_text(_extract_cells(row, header, "customer_name", "name"), max_len=200)
        if not name:
            continue
        code = _clean_text(_extract_cells(row, header, "account_code", "acc code", "account code", "code"), max_len=40)
        contact = _clean_text(_extract_cells(row, header, "billing contact", "contact"), max_len=160)
        terms = _clean_text(_extract_cells(row, header, "payment_terms", "terms"), max_len=80)
        customer_type = _clean_text(_extract_cells(row, header, "customer_type"), max_len=60)
        operation_area = _clean_text(_extract_cells(row, header, "operation_area"), max_len=120)
        source_file = _clean_text(_extract_cells(row, header, "source_file"), max_len=255) or path.path.name
        _row, created = _upsert_account_customer(
            db,
            club_id=int(club_id),
            name=name,
            account_code=code,
            billing_contact=contact,
            terms_label=terms,
            customer_type=customer_type,
            operation_area=operation_area,
            source_file=source_file,
            import_reference=f"{source_file}:{code or name}",
        )
        if getattr(_row, "import_reference", None):
            touched_refs.add(str(_row.import_reference))
        if created:
            inserted += 1
        else:
            updated += 1
    return {"rows": max(0, len(rows) - header_idx - 1), "inserted": inserted, "updated": updated, "_touched_refs": sorted(touched_refs)}


def _find_member_row(
    db: Session,
    *,
    club_id: int,
    member_number: str | None,
    email: str | None,
    first_name: str,
    last_name: str,
) -> models.Member | None:
    if member_number:
        row = (
            db.query(models.Member)
            .filter(models.Member.club_id == int(club_id), models.Member.member_number == member_number)
            .first()
        )
        if row:
            return row
    if email:
        row = (
            db.query(models.Member)
            .filter(models.Member.club_id == int(club_id), func.lower(models.Member.email) == email.lower())
            .first()
        )
        if row:
            return row
    return (
        db.query(models.Member)
        .filter(
            models.Member.club_id == int(club_id),
            func.lower(models.Member.first_name) == first_name.lower(),
            func.lower(models.Member.last_name) == last_name.lower(),
        )
        .order_by(models.Member.id.asc())
        .first()
    )


def _member_record_score(row: models.Member) -> int:
    score = 0
    if str(getattr(row, "member_number", "") or "").strip():
        score += 100
    if str(getattr(row, "import_reference", "") or "").strip():
        score += 40
    if str(getattr(row, "email", "") or "").strip():
        score += 40
    if str(getattr(row, "phone", "") or "").strip():
        score += 20
    if getattr(row, "person_id", None):
        score += 80
    if str(getattr(row, "handicap_number", "") or "").strip():
        score += 10
    if str(getattr(row, "home_club", "") or "").strip():
        score += 5
    if getattr(row, "membership_date", None):
        score += 5
    if getattr(row, "membership_expiration", None):
        score += 5
    return score


def _member_duplicate_conflict(canonical: models.Member, duplicate: models.Member) -> bool:
    comparable_fields = (
        "member_number",
        "email",
        "phone",
        "person_id",
        "membership_status",
        "membership_date",
        "membership_expiration",
    )
    for field_name in comparable_fields:
        left = getattr(canonical, field_name, None)
        right = getattr(duplicate, field_name, None)
        if left in (None, "") or right in (None, ""):
            continue
        if left != right:
            return True
    return False


def _merge_member_rows(canonical: models.Member, duplicate: models.Member) -> None:
    for field_name in (
        "member_number",
        "email",
        "phone",
        "handicap_number",
        "home_club",
        "country_of_residence",
        "membership_category",
        "membership_category_raw",
        "primary_operation",
        "membership_status",
        "member_lifecycle_status",
        "record_status",
        "person_type",
        "membership_date",
        "membership_expiration",
        "source_file",
        "source_row_number",
        "import_reference",
        "golf_access",
        "tennis_access",
        "bowls_access",
        "squash_access",
        "gender",
        "player_category",
        "student",
        "handicap_index",
        "handicap_sa_id",
        "person_id",
    ):
        current = getattr(canonical, field_name, None)
        incoming = getattr(duplicate, field_name, None)
        if current in (None, "") and incoming not in (None, ""):
            setattr(canonical, field_name, incoming)
    if int(getattr(duplicate, "active", 0) or 0) == 1:
        canonical.active = 1


def dedupe_umhlali_members(db: Session, *, club_id: int) -> dict[str, int]:
    rows = (
        db.query(models.Member)
        .filter(models.Member.club_id == int(club_id))
        .order_by(models.Member.id.asc())
        .all()
    )
    groups: dict[tuple[str, str, str, str, str, str], list[models.Member]] = {}
    for row in rows:
        key = member_identity_key(
            first_name=getattr(row, "first_name", None),
            last_name=getattr(row, "last_name", None),
            membership_category=getattr(row, "membership_category", None),
            membership_status=getattr(row, "membership_status", None),
            membership_date=getattr(row, "membership_date", None),
            membership_expiration=getattr(row, "membership_expiration", None),
        )
        groups.setdefault(key, []).append(row)

    duplicate_groups = 0
    merged_rows = 0
    skipped_groups = 0
    booking_relinks = 0
    flush_interval = 100
    pending_writes = 0

    for group_rows in groups.values():
        if len(group_rows) <= 1:
            continue
        duplicate_groups += 1
        ordered = sorted(group_rows, key=lambda row: (-_member_record_score(row), int(getattr(row, "id", 0) or 0)))
        canonical = ordered[0]
        duplicates = ordered[1:]

        safe = True
        for duplicate in duplicates:
            if _member_duplicate_conflict(canonical, duplicate):
                safe = False
                break
            if str(getattr(duplicate, "member_number", "") or "").strip() and str(
                getattr(canonical, "member_number", "") or ""
            ).strip():
                safe = False
                break
            if getattr(duplicate, "person_id", None) and getattr(canonical, "person_id", None):
                safe = False
                break
        if not safe:
            skipped_groups += 1
            continue

        duplicate_ids = [int(duplicate.id) for duplicate in duplicates if getattr(duplicate, "id", None)]
        booking_counts = {}
        if duplicate_ids:
            booking_counts = {
                int(member_id): int(count or 0)
                for member_id, count in (
                    db.query(models.Booking.member_id, func.count(models.Booking.id))
                    .filter(models.Booking.member_id.in_(duplicate_ids))
                    .group_by(models.Booking.member_id)
                    .all()
                )
            }

        for duplicate in duplicates:
            duplicate_id = int(duplicate.id)
            if int(booking_counts.get(duplicate_id, 0) or 0) > 0:
                booking_relinks += int(
                    db.query(models.Booking)
                    .filter(models.Booking.member_id == duplicate_id)
                    .update({models.Booking.member_id: int(canonical.id)}, synchronize_session=False)
                    or 0
                )
            _merge_member_rows(canonical, duplicate)
            db.delete(duplicate)
            merged_rows += 1
            pending_writes += 1
            if pending_writes >= flush_interval:
                db.flush()
                pending_writes = 0

    if pending_writes > 0:
        db.flush()

    return {
        "duplicate_groups": int(duplicate_groups),
        "merged_rows": int(merged_rows),
        "skipped_groups": int(skipped_groups),
        "booking_relinks": int(booking_relinks),
    }


def _ingest_members(
    db: Session,
    *,
    club_id: int,
    path: SheetRef,
) -> dict[str, int]:
    rows = _sheet_rows(path)
    if not rows:
        return {"rows": 0, "inserted": 0, "updated": 0, "people_linked": 0}

    header_info = _find_members_header(rows)
    if header_info is None:
        return {"rows": 0, "inserted": 0, "updated": 0, "people_linked": 0}
    header_idx, header = header_info
    inserted = 0
    updated = 0
    existing_rows = db.query(models.Member).filter(models.Member.club_id == int(club_id)).all()
    by_member_number = {
        str(getattr(row, "member_number", "") or "").strip(): row
        for row in existing_rows
        if str(getattr(row, "member_number", "") or "").strip()
    }
    by_email = {
        str(getattr(row, "email", "") or "").strip().lower(): row
        for row in existing_rows
        if str(getattr(row, "email", "") or "").strip()
    }
    by_identity = {}
    for row in existing_rows:
        key = member_identity_key(
            first_name=getattr(row, "first_name", None),
            last_name=getattr(row, "last_name", None),
            membership_category=getattr(row, "membership_category", None),
            membership_status=getattr(row, "membership_status", None),
            membership_date=getattr(row, "membership_date", None),
            membership_expiration=getattr(row, "membership_expiration", None),
        )
        by_identity[key] = row
    by_import_reference = {
        str(getattr(row, "import_reference", "") or "").strip(): row
        for row in existing_rows
        if str(getattr(row, "import_reference", "") or "").strip()
    }
    pending_by_member_number: dict[str, dict[str, Any]] = {}
    pending_by_email: dict[str, dict[str, Any]] = {}
    pending_by_identity: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    pending_by_import_reference: dict[str, dict[str, Any]] = {}
    insert_payloads: list[dict[str, Any]] = []
    touched_refs: set[str] = set()

    for row in rows[header_idx + 1 :]:
        first = _clean_text(_extract_cells(row, header, "first_name", "first name"), max_len=120)
        last = _clean_text(_extract_cells(row, header, "last_name", "last name"), max_len=120)
        if not first or not last:
            continue

        source_row_raw = _extract_cells(row, header, "source_row_number", "source row number")
        try:
            source_row_number = int(str(source_row_raw or "").strip()) if str(source_row_raw or "").strip() else None
        except Exception:
            source_row_number = None
        member_number = _clean_text(
            _extract_cells(row, header, "member_number", "member no", "member number"),
            max_len=50,
        )
        email = _clean_text(_extract_cells(row, header, "email", "member_email"), max_len=200)
        if email:
            email = email.lower()
        phone = _clean_text(_extract_cells(row, header, "phone", "mobile", "cell"), max_len=50)
        country = _clean_text(_extract_cells(row, header, "country_of_residence", "country of residence"), max_len=120)
        home_club = _clean_text(_extract_cells(row, header, "home_club", "home club"), max_len=120) or "Umhlali Country Club"
        membership = _clean_text(
            _extract_cells(row, header, "membership_category_raw", "membership_category", "membership"),
            max_len=160,
        ) or "Unspecified"
        primary_operation = normalize_primary_operation(
            _extract_cells(row, header, "primary_operation", "operation_area"),
            membership,
        )
        membership_date = parse_membership_date(
            _extract_cells(row, header, "membership_start_date", "membership date", "membership_date")
        )
        membership_exp = parse_membership_date(
            _extract_cells(row, header, "membership_expiration_date", "membership expiration", "membership_expiration")
        )
        status_raw = _clean_text(_extract_cells(row, header, "member_lifecycle_status", "status"), max_len=40) or "Active"
        norm_status = normalize_membership_status(status_raw)
        active_flag = 1 if norm_status == "active" else 0
        record_status = _clean_text(_extract_cells(row, header, "record_status"), max_len=40) or norm_status
        person_type = _clean_text(_extract_cells(row, header, "person_type"), max_len=40) or "Member"
        gender = _clean_text(_extract_cells(row, header, "gender"), max_len=20)
        source_file = _clean_text(_extract_cells(row, header, "source_file"), max_len=255) or path.path.name
        import_reference = f"{source_file}:{source_row_number}" if source_row_number is not None else None
        payload = {
            "club_id": int(club_id),
            "member_number": member_number,
            "first_name": first,
            "last_name": last,
            "email": email,
            "phone": phone,
            "home_club": home_club,
            "country_of_residence": country,
            "membership_category": membership,
            "membership_category_raw": membership,
            "primary_operation": primary_operation,
            "membership_status": norm_status,
            "member_lifecycle_status": norm_status,
            "record_status": record_status,
            "person_type": person_type,
            "membership_date": membership_date,
            "membership_expiration": membership_exp,
            "source_file": source_file,
            "source_row_number": source_row_number,
            "import_reference": import_reference,
            "golf_access": parse_yes_no_flag(_extract_cells(row, header, "golf_access")),
            "tennis_access": parse_yes_no_flag(_extract_cells(row, header, "tennis_access")),
            "bowls_access": parse_yes_no_flag(_extract_cells(row, header, "bowls_access")),
            "squash_access": parse_yes_no_flag(_extract_cells(row, header, "squash_access")),
            "active": active_flag,
            "gender": gender,
            "player_category": classify_membership_group(primary_operation or membership),
        }
        identity_key = member_identity_key(
            first_name=first,
            last_name=last,
            membership_category=membership,
            membership_status=norm_status,
            membership_date=membership_date,
            membership_expiration=membership_exp,
        )

        existing = None
        if import_reference:
            existing = by_import_reference.get(import_reference)
        if existing is None and import_reference:
            existing = pending_by_import_reference.get(import_reference)
        if member_number:
            existing = existing or by_member_number.get(member_number)
        if existing is None and member_number:
            existing = pending_by_member_number.get(member_number)
        if existing is None and email:
            existing = by_email.get(email)
        if existing is None and email:
            existing = pending_by_email.get(email)
        if existing is None:
            existing = by_identity.get(identity_key)
        if existing is None:
            existing = pending_by_identity.get(identity_key)
        if existing is None:
            insert_payloads.append(payload)
            inserted += 1
        else:
            if isinstance(existing, dict):
                existing.update(payload)
            else:
                if member_number:
                    existing.member_number = member_number
                existing.first_name = first
                existing.last_name = last
                existing.email = email or existing.email
                existing.phone = phone or existing.phone
                existing.home_club = home_club or existing.home_club or "Umhlali Country Club"
                existing.country_of_residence = country
                existing.membership_category = membership
                existing.membership_category_raw = membership
                existing.primary_operation = primary_operation
                existing.membership_status = norm_status
                existing.member_lifecycle_status = norm_status
                existing.record_status = record_status
                existing.person_type = person_type
                existing.membership_date = membership_date
                existing.membership_expiration = membership_exp
                existing.source_file = source_file
                existing.source_row_number = source_row_number
                existing.import_reference = import_reference
                existing.golf_access = payload["golf_access"]
                existing.tennis_access = payload["tennis_access"]
                existing.bowls_access = payload["bowls_access"]
                existing.squash_access = payload["squash_access"]
                existing.active = active_flag
                existing.gender = gender or existing.gender
                existing.player_category = classify_membership_group(primary_operation or membership)
                updated += 1
        if import_reference:
            touched_refs.add(import_reference)
        if import_reference:
            if isinstance(existing, dict):
                pending_by_import_reference[import_reference] = existing
            else:
                by_import_reference[import_reference] = existing
        if member_number:
            if isinstance(existing, dict):
                pending_by_member_number[member_number] = existing
            else:
                by_member_number[member_number] = existing
        if email:
            if isinstance(existing, dict):
                pending_by_email[email] = existing
            else:
                by_email[email] = existing
        if isinstance(existing, dict):
            pending_by_identity[identity_key] = existing
        else:
            by_identity[identity_key] = existing

    if insert_payloads:
        chunk_size = 100
        for start in range(0, len(insert_payloads), chunk_size):
            db.bulk_insert_mappings(models.Member, insert_payloads[start : start + chunk_size])
            db.flush()
    return {
        "rows": max(0, len(rows) - header_idx - 1),
        "inserted": inserted,
        "updated": updated,
        "people_linked": 0,
        "_touched_refs": sorted(touched_refs),
    }


def _upsert_staff_role_profile(
    db: Session,
    *,
    club_id: int,
    staff_name: str,
    role_label: str,
    operation_area: str | None = None,
    user_type: str | None = None,
    source_file: str | None = None,
) -> tuple[models.StaffRoleProfile, bool]:
    row = (
        db.query(models.StaffRoleProfile)
        .filter(
            models.StaffRoleProfile.club_id == int(club_id),
            func.lower(models.StaffRoleProfile.staff_name) == staff_name.lower(),
            func.lower(models.StaffRoleProfile.role_label) == role_label.lower(),
        )
        .first()
    )
    linked_user = (
        db.query(models.User)
        .filter(
            models.User.club_id == int(club_id),
            models.User.role.in_([models.UserRole.admin, models.UserRole.club_staff]),
            func.lower(models.User.name) == staff_name.lower(),
        )
        .first()
    )
    linked_user_id = int(linked_user.id) if linked_user else None

    created = False
    if row is None:
        row = models.StaffRoleProfile(
            club_id=int(club_id),
            staff_name=staff_name,
            role_label=role_label,
            linked_user_id=linked_user_id,
            operation_area=operation_area,
            user_type=user_type,
            source_file=source_file,
            active=1,
        )
        db.add(row)
        db.flush()
        created = True
    else:
        row.linked_user_id = linked_user_id
        row.operation_area = operation_area
        row.user_type = user_type
        row.source_file = source_file
        row.active = 1
        row.updated_at = datetime.utcnow()
    return row, created


def _ingest_staff_roles(
    db: Session,
    *,
    club_id: int,
    path: SheetRef,
) -> dict[str, int]:
    rows = _sheet_rows(path)
    if not rows:
        return {"rows": 0, "inserted": 0, "updated": 0, "linked_users": 0}
    header_info = _find_header_row(rows, "staff_name", "role") or _find_header_row(rows, "name", "role")
    if header_info is None:
        return {"rows": 0, "inserted": 0, "updated": 0, "linked_users": 0}
    header_idx, header = header_info
    inserted = 0
    updated = 0
    linked_users = 0
    touched_refs: set[str] = set()
    for row in rows[header_idx + 1 :]:
        staff_name = _clean_text(_extract_cells(row, header, "staff_name", "name"), max_len=160)
        role_label = _clean_text(_extract_cells(row, header, "role"), max_len=120)
        if not staff_name or not role_label:
            continue
        operation_area = _clean_text(_extract_cells(row, header, "operation_area"), max_len=120)
        user_type = _clean_text(_extract_cells(row, header, "user_type"), max_len=60)
        source_file = _clean_text(_extract_cells(row, header, "source_file"), max_len=255) or path.path.name
        saved, created = _upsert_staff_role_profile(
            db,
            club_id=int(club_id),
            staff_name=staff_name,
            role_label=role_label,
            operation_area=operation_area,
            user_type=user_type,
            source_file=source_file,
        )
        if created:
            inserted += 1
        else:
            updated += 1
        if getattr(saved, "linked_user_id", None):
            linked_users += 1
        touched_refs.add(f"{staff_name.lower()}|{role_label.lower()}")
    return {
        "rows": max(0, len(rows) - header_idx - 1),
        "inserted": inserted,
        "updated": updated,
        "linked_users": linked_users,
        "_touched_refs": sorted(touched_refs),
    }


def _find_golf_day_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows):
        cells = [str(v or "").strip().lower() for v in row]
        if ("booking_name" in cells and "invoice_number" in cells) or ("name" in cells and ("date of golf day" in cells or "amount" in cells)):
            return idx, _header_index_map(row)
    return None


def _parse_amount_and_date(value: Any) -> tuple[float | None, date | None, str | None]:
    raw = _clean_text(value, max_len=200)
    if not raw:
        return None, None, None
    amount = _to_float(raw)
    parsed_date = _to_date(raw)
    return amount, parsed_date, raw


def _resolve_account_customer(
    db: Session,
    *,
    club_id: int,
    account_code: str | None,
    event_name: str,
) -> models.AccountCustomer | None:
    if account_code:
        row = (
            db.query(models.AccountCustomer)
            .filter(
                models.AccountCustomer.club_id == int(club_id),
                func.lower(models.AccountCustomer.account_code) == account_code.lower(),
            )
            .first()
        )
        if row:
            return row
    token = event_name.strip().lower()
    if not token:
        return None
    exact = (
        db.query(models.AccountCustomer)
        .filter(
            models.AccountCustomer.club_id == int(club_id),
            func.lower(models.AccountCustomer.name) == token,
        )
        .first()
    )
    if exact:
        return exact

    candidates = (
        db.query(models.AccountCustomer)
        .filter(models.AccountCustomer.club_id == int(club_id))
        .order_by(models.AccountCustomer.id.asc())
        .all()
    )
    for row in candidates:
        name = str(getattr(row, "name", "") or "").strip().lower()
        if not name:
            continue
        if name in token or token in name:
            return row
    return None


def _normalize_payment_status(
    *,
    amount: float | None,
    balance_due: float | None,
    deposit_amount: float | None,
    full_payment_amount: float | None,
) -> str:
    full = float(full_payment_amount or 0.0)
    total = float(amount or 0.0)
    balance = float(balance_due or 0.0)
    deposit = float(deposit_amount or 0.0)
    if full > 0:
        return "paid"
    if total > 0 and balance <= 0:
        return "paid"
    if deposit > 0:
        return "partial"
    if total > 0 and 0 < balance < total:
        return "partial"
    return "pending"


def _ingest_golf_day_bookings(
    db: Session,
    *,
    club_id: int,
    path: SheetRef,
) -> dict[str, int]:
    rows = _sheet_rows(path)
    if not rows:
        return {"rows": 0, "inserted": 0, "updated": 0}

    header_info = _find_golf_day_header(rows)
    if header_info is None:
        return {"rows": 0, "inserted": 0, "updated": 0}
    header_idx, header = header_info
    inserted = 0
    updated = 0
    touched_refs: set[str] = set()

    for row in rows[header_idx + 1 :]:
        event_name = _clean_text(_extract_cells(row, header, "booking_name", "name"), max_len=220)
        if not event_name:
            continue
        event_date_raw = _clean_text(_extract_cells(row, header, "event_date_raw", "date of golf day"), max_len=120)
        event_date = _to_date(_extract_cells(row, header, "event_start_date")) or _to_date(event_date_raw)
        event_end_date = _to_date(_extract_cells(row, header, "event_end_date"))
        amount = _to_float(_extract_cells(row, header, "gross_amount_zar", "amount"))
        invoice_reference = _clean_text(_extract_cells(row, header, "invoice_number", "invoiced"), max_len=80)

        deposit_amount, deposit_date, deposit_note = _parse_amount_and_date(
            _extract_cells(row, header, "deposit_received_raw", "deposit received & date")
        )
        balance_due = _to_float(_extract_cells(row, header, "balance_due_zar", "balance due"))
        full_payment_amount, full_payment_date, full_payment_note = _parse_amount_and_date(
            _extract_cells(row, header, "full_payment_received_raw", "full payment received & date")
        )
        notes = _clean_text(_extract_cells(row, header, "notes"), max_len=2000)
        operation_area = _clean_text(_extract_cells(row, header, "operation_area"), max_len=120)
        source_file = _clean_text(_extract_cells(row, header, "source_file"), max_len=255) or path.path.name
        import_reference = invoice_reference or f"{source_file}:{event_name}:{event_date_raw or event_date or ''}"

        account_code_guess = None
        customer = _resolve_account_customer(
            db,
            club_id=int(club_id),
            account_code=account_code_guess,
            event_name=event_name,
        )
        account_customer_id = int(customer.id) if customer else None
        account_code_snapshot = str(getattr(customer, "account_code", "") or "").strip() or None
        contact_name = str(getattr(customer, "billing_contact", "") or "").strip() or None
        payment_status = _normalize_payment_status(
            amount=amount,
            balance_due=balance_due,
            deposit_amount=deposit_amount,
            full_payment_amount=full_payment_amount,
        )

        existing = (
            db.query(models.GolfDayBooking)
            .filter(
                models.GolfDayBooking.club_id == int(club_id),
                func.lower(models.GolfDayBooking.event_name) == event_name.lower(),
                models.GolfDayBooking.event_date_raw == event_date_raw,
                models.GolfDayBooking.invoice_reference == invoice_reference,
            )
            .first()
        )

        if existing is None:
            existing = models.GolfDayBooking(
                club_id=int(club_id),
                account_customer_id=account_customer_id,
                event_name=event_name,
                event_date=event_date,
                event_end_date=event_end_date,
                event_date_raw=event_date_raw,
                amount=float(amount or 0.0),
                invoice_reference=invoice_reference,
                deposit_amount=deposit_amount,
                deposit_received_date=deposit_date,
                deposit_received_note=deposit_note,
                balance_due=balance_due,
                full_payment_amount=full_payment_amount,
                full_payment_date=full_payment_date,
                full_payment_note=full_payment_note,
                payment_status=payment_status,
                contact_name=contact_name,
                account_code_snapshot=account_code_snapshot,
                operation_area=operation_area,
                source_file=source_file,
                import_reference=import_reference,
                notes=notes,
            )
            db.add(existing)
            db.flush()
            inserted += 1
        else:
            existing.account_customer_id = account_customer_id
            existing.event_date = event_date
            existing.event_end_date = event_end_date
            existing.amount = float(amount or 0.0)
            existing.deposit_amount = deposit_amount
            existing.deposit_received_date = deposit_date
            existing.deposit_received_note = deposit_note
            existing.balance_due = balance_due
            existing.full_payment_amount = full_payment_amount
            existing.full_payment_date = full_payment_date
            existing.full_payment_note = full_payment_note
            existing.payment_status = payment_status
            existing.contact_name = contact_name
            existing.account_code_snapshot = account_code_snapshot
            existing.operation_area = operation_area
            existing.source_file = source_file
            existing.import_reference = import_reference
            existing.notes = notes
            existing.updated_at = datetime.utcnow()
            updated += 1
        if import_reference:
            touched_refs.add(import_reference)

    return {"rows": max(0, len(rows) - header_idx - 1), "inserted": inserted, "updated": updated, "_touched_refs": sorted(touched_refs)}


def _purge_force_reload_base(db: Session, *, club_id: int) -> dict[str, int]:
    golf_day_deleted = int(
        db.query(models.GolfDayBooking)
        .filter(models.GolfDayBooking.club_id == int(club_id))
        .delete(synchronize_session=False)
        or 0
    )
    staff_deleted = int(
        db.query(models.StaffRoleProfile)
        .filter(models.StaffRoleProfile.club_id == int(club_id))
        .delete(synchronize_session=False)
        or 0
    )
    db.flush()
    return {
        "golf_day_deleted": golf_day_deleted,
        "staff_roles_deleted": staff_deleted,
    }


def _cleanup_forced_member_reload(
    db: Session,
    *,
    club_id: int,
    touched_refs: set[str],
) -> dict[str, int]:
    rows = db.query(models.Member).filter(models.Member.club_id == int(club_id)).all()
    stale = [row for row in rows if str(getattr(row, "import_reference", "") or "").strip() not in touched_refs]
    if not stale:
        return {"deleted": 0, "archived": 0}

    stale_ids = [int(row.id) for row in stale if getattr(row, "id", None)]
    booking_counts: dict[int, int] = {}
    if stale_ids:
        booking_counts = {
            int(member_id): int(count or 0)
            for member_id, count in (
                db.query(models.Booking.member_id, func.count(models.Booking.id))
                .filter(models.Booking.member_id.in_(stale_ids))
                .group_by(models.Booking.member_id)
                .all()
            )
        }

    deleted = 0
    archived = 0
    for row in stale:
        if int(booking_counts.get(int(row.id), 0) or 0) > 0:
            row.active = 0
            row.membership_status = "inactive"
            row.member_lifecycle_status = "inactive"
            row.record_status = row.record_status or "archived_after_force_reload"
            row.updated_at = datetime.utcnow()
            archived += 1
        else:
            db.delete(row)
            deleted += 1
    db.flush()
    return {"deleted": deleted, "archived": archived}


def _cleanup_forced_account_reload(
    db: Session,
    *,
    club_id: int,
    touched_refs: set[str],
) -> dict[str, int]:
    rows = db.query(models.AccountCustomer).filter(models.AccountCustomer.club_id == int(club_id)).all()
    stale = [row for row in rows if str(getattr(row, "import_reference", "") or "").strip() not in touched_refs]
    if not stale:
        return {"deleted": 0, "archived": 0}

    stale_ids = [int(row.id) for row in stale if getattr(row, "id", None)]
    booking_counts: dict[int, int] = {}
    if stale_ids:
        booking_counts = {
            int(account_id): int(count or 0)
            for account_id, count in (
                db.query(models.Booking.account_customer_id, func.count(models.Booking.id))
                .filter(models.Booking.account_customer_id.in_(stale_ids))
                .group_by(models.Booking.account_customer_id)
                .all()
            )
        }

    deleted = 0
    archived = 0
    for row in stale:
        if int(booking_counts.get(int(row.id), 0) or 0) > 0:
            row.active = 0
            existing_note = str(getattr(row, "notes", "") or "").strip()
            suffix = "Archived after Umhlali force reload."
            row.notes = f"{existing_note} {suffix}".strip()
            row.updated_at = datetime.utcnow()
            archived += 1
        else:
            db.delete(row)
            deleted += 1
    db.flush()
    return {"deleted": deleted, "archived": archived}


def seed_umhlali_operational_inputs(
    db: Session,
    *,
    club_id: int,
    force: bool = False,
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "status": "skipped",
        "setup_dir": None,
        "missing_files": [],
        "errors": [],
        "force_cleanup": {},
        "members": {},
        "accounts": {},
        "golf_day": {},
        "staff_roles": {},
    }

    setup = _find_setup_files()
    if setup is None:
        out["status"] = "missing_inputs"
        out["missing_files"] = [
            "Member List -Golf Scape ...xlsx",
            "Account Customers.xlsx",
            "Golf Day Bookings.xlsx",
            "Staff User List.xlsx",
        ]
        return out

    out["setup_dir"] = str(setup.setup_dir)

    if not force:
        dedup_stats = dedupe_umhlali_members(db, club_id=int(club_id))
        if int(dedup_stats.get("merged_rows", 0) or 0) > 0:
            out["members"] = {"deduped_rows": int(dedup_stats.get("merged_rows", 0))}
        has_members = bool(
            db.query(models.Member.id).filter(models.Member.club_id == int(club_id)).first()
        )
        has_accounts = bool(
            db.query(models.AccountCustomer.id).filter(models.AccountCustomer.club_id == int(club_id)).first()
        )
        has_golf_day = bool(
            db.query(models.GolfDayBooking.id).filter(models.GolfDayBooking.club_id == int(club_id)).first()
        )
        has_staff_roles = bool(
            db.query(models.StaffRoleProfile.id).filter(models.StaffRoleProfile.club_id == int(club_id)).first()
        )
        if has_members and has_accounts and has_golf_day and has_staff_roles:
            out["status"] = "already_seeded"
            return out

    sources_loaded = 0
    touched_refs: dict[str, set[str]] = {
        "members": set(),
        "accounts": set(),
        "golf_day": set(),
        "staff_roles": set(),
    }

    if force:
        out["force_cleanup"] = _purge_force_reload_base(db, club_id=int(club_id))

    def _load_source(
        *,
        output_key: str,
        source_name: str,
        missing_label: str,
        path: SheetRef | None,
        loader,
    ) -> None:
        nonlocal sources_loaded
        if path is None or not path.path.exists():
            out["missing_files"].append(missing_label)
            return
        try:
            with db.begin_nested():
                batch = _upsert_import_batch(
                    db,
                    club_id=int(club_id),
                    source=source_name,
                    file_name=path.path.name,
                )
                stats = loader(db, club_id=int(club_id), path=path)
                raw_refs = stats.pop("_touched_refs", [])
                touched_refs[output_key] = {
                    str(value).strip()
                    for value in raw_refs
                    if str(value or "").strip()
                }
                batch.rows_total = int(stats.get("rows", 0))
                batch.rows_inserted = int(stats.get("inserted", 0))
                batch.rows_updated = int(stats.get("updated", 0))
                batch.rows_failed = 0
                out[output_key] = stats
            sources_loaded += 1
        except Exception as exc:
            out["errors"].append(f"{source_name}: {type(exc).__name__}: {str(exc)[:180]}")
            out[output_key] = {"status": "failed"}

    _load_source(
        output_key="members",
        source_name="umhlali_members_xlsx",
        missing_label="Member List -Golf Scape ...xlsx",
        path=setup.member_list,
        loader=_ingest_members,
    )
    _load_source(
        output_key="accounts",
        source_name="umhlali_account_customers_xlsx",
        missing_label="Account Customers.xlsx",
        path=setup.account_customers,
        loader=_ingest_account_customers,
    )
    _load_source(
        output_key="staff_roles",
        source_name="umhlali_staff_roles_xlsx",
        missing_label="Staff User List.xlsx",
        path=setup.staff_roles,
        loader=_ingest_staff_roles,
    )
    _load_source(
        output_key="golf_day",
        source_name="umhlali_golf_day_bookings_xlsx",
        missing_label="Golf Day Bookings.xlsx",
        path=setup.golf_day_bookings,
        loader=_ingest_golf_day_bookings,
    )

    if force and touched_refs["members"]:
        out["force_cleanup"]["members"] = _cleanup_forced_member_reload(
            db,
            club_id=int(club_id),
            touched_refs=touched_refs["members"],
        )
    if force and touched_refs["accounts"]:
        out["force_cleanup"]["accounts"] = _cleanup_forced_account_reload(
            db,
            club_id=int(club_id),
            touched_refs=touched_refs["accounts"],
        )

    dedup_stats = dedupe_umhlali_members(db, club_id=int(club_id))
    if int(dedup_stats.get("merged_rows", 0) or 0) > 0:
        current = out.get("members") if isinstance(out.get("members"), dict) else {}
        current["deduped_rows"] = int(dedup_stats.get("merged_rows", 0))
        out["members"] = current

    out["status"] = "seeded" if sources_loaded > 0 else "missing_inputs"
    return out
